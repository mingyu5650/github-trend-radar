#!/usr/bin/env python3
"""Tests for the Excel-backed project watchlist."""

import copy
import fcntl
import os
import multiprocessing
import posixpath
import re
import stat
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

from storage import lock_path_for  # noqa: E402
from watchlist import (  # noqa: E402
    AUTOMATIC_FIELDS,
    HEADERS,
    MANUAL_FIELDS,
    SHEET_NAME,
    WatchlistError,
    create_watchlist,
    read_watchlist,
    update_automatic_fields,
)


def _append_row(path, values):
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    worksheet.append(values)
    workbook.save(path)
    workbook.close()


def _concurrent_update(path, start_event, repository, version):
    """Process target kept at module scope for multiprocessing compatibility."""
    start_event.wait()
    update_automatic_fields(path, {repository: {"最新版本": version}})


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _worksheet_entry(path, sheet_name=SHEET_NAME):
    with zipfile.ZipFile(path, "r") as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        relationship_root = ET.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
    relationship_id = None
    for sheet in workbook_root.findall(".//{{{}}}sheet".format(_MAIN_NS)):
        if sheet.get("name") == sheet_name:
            relationship_id = sheet.get("{{{}}}id".format(_DOC_REL_NS))
            break
    if relationship_id is None:
        raise AssertionError("worksheet relationship is missing")
    for relationship in relationship_root.findall(
        "{{{}}}Relationship".format(_PACKAGE_REL_NS)
    ):
        if relationship.get("Id") == relationship_id:
            target = relationship.get("Target")
            if target.startswith("/"):
                return target.lstrip("/")
            return posixpath.normpath(posixpath.join("xl", target))
    raise AssertionError("worksheet target is missing")


def _zip_payloads(path):
    with zipfile.ZipFile(path, "r") as archive:
        return {info.filename: archive.read(info) for info in archive.infolist()}


def _worksheet_cells(xml_bytes, columns=range(1, 13)):
    root = ET.fromstring(xml_bytes)
    result = {}
    allowed = set(columns)
    for cell in root.findall(".//{{{}}}c".format(_MAIN_NS)):
        reference = cell.get("r", "")
        letters = "".join(character for character in reference if character.isalpha())
        number = 0
        for character in letters.upper():
            number = number * 26 + ord(character) - ord("A") + 1
        if number in allowed:
            result[reference] = ET.tostring(cell, encoding="utf-8")
    return result


def _inject_formula_cache(path, sheet_entry, cell_reference, cached_value):
    payloads = _zip_payloads(path)
    root = ET.fromstring(payloads[sheet_entry])
    cell = next(
        item
        for item in root.findall(".//{{{}}}c".format(_MAIN_NS))
        if item.get("r") == cell_reference
    )
    value = cell.find("{{{}}}v".format(_MAIN_NS))
    if value is None:
        value = ET.SubElement(cell, "{{{}}}v".format(_MAIN_NS))
    value.text = cached_value
    payloads[sheet_entry] = ET.tostring(
        root, encoding="utf-8", xml_declaration=True
    )
    temp_path = path.with_name("injected.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp_path, "w"
    ) as destination:
        for info in source.infolist():
            destination.writestr(info, payloads[info.filename])
    os.replace(temp_path, path)


def _rewrite_zip_payload(path, entry_name, transform):
    temp_path = path.with_name("rewritten.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temp_path, "w"
    ) as destination:
        destination.comment = source.comment
        for info in source.infolist():
            payload = source.read(info)
            if info.filename == entry_name:
                payload = transform(payload)
            destination.writestr(info, payload)
    os.replace(temp_path, path)


class WatchlistTemplateTests(unittest.TestCase):
    def test_creates_readable_template_with_fixed_columns_and_visual_structure(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "nested" / "watchlist.xlsx"

            create_watchlist(path)

            self.assertTrue(path.exists())
            workbook = load_workbook(path)
            worksheet = workbook[SHEET_NAME]
            self.assertEqual(
                tuple(cell.value for cell in worksheet[1]), tuple(HEADERS)
            )
            self.assertEqual(len(HEADERS), 16)
            self.assertEqual(tuple(HEADERS[:12]), tuple(MANUAL_FIELDS))
            self.assertEqual(tuple(HEADERS[12:]), tuple(AUTOMATIC_FIELDS))
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, "A1:P1")

            for cell in worksheet[1][:12]:
                self.assertTrue(cell.font.bold)
                self.assertEqual(cell.font.color.type, "rgb")
                self.assertEqual(cell.font.color.rgb, "FFFFFFFF")
                self.assertEqual(cell.fill.fgColor.rgb, "FF4472C4")
            for cell in worksheet[1][12:]:
                self.assertTrue(cell.font.bold)
                self.assertEqual(cell.font.color.rgb, "FFFFFFFF")
                self.assertEqual(cell.fill.fgColor.rgb, "FF70AD47")

            validations = {
                str(validation.sqref): validation.formula1
                for validation in worksheet.data_validations.dataValidation
            }
            self.assertIn("C2:C1000", validations)
            self.assertIn("D2:D1000", validations)
            self.assertIn("I2:I1000", validations)
            self.assertIn("J2:J1000", validations)
            self.assertGreater(worksheet.column_dimensions["A"].width, 10)
            self.assertGreater(worksheet.column_dimensions["E"].width, 10)
            self.assertEqual(worksheet.max_row, 1)
            self.assertEqual(worksheet.max_column, 16)
            workbook.close()

    def test_existing_valid_file_is_validated_without_being_overwritten(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["ORG/Tool", "项目", "试用中"])
            original = path.read_bytes()

            create_watchlist(path)

            self.assertEqual(path.read_bytes(), original)

    def test_existing_bad_structure_is_rejected_without_replacement(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            workbook = load_workbook(path)
            workbook[SHEET_NAME]["B1"] = "私密错误表头"
            workbook.save(path)
            workbook.close()
            original = path.read_bytes()

            with self.assertRaises(WatchlistError) as caught:
                create_watchlist(path)

            self.assertNotIn("私密错误表头", str(caught.exception))
            self.assertEqual(path.read_bytes(), original)

    def test_missing_file_is_created_and_empty_watchlist_reads_as_empty(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"

            self.assertEqual(read_watchlist(path), [])

            self.assertTrue(path.exists())


class WatchlistReadTests(unittest.TestCase):
    def test_reads_unicode_rows_skips_empty_rows_and_normalizes_only_repository(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, [None] * 16)
            values = [
                " OpenAI/GPT-OSS ",
                "中文项目 ",
                "试用中",
                "高",
                "本地 AI 工具",
                "v1",
                "2026-07-14",
                "替代项目",
                "是",
                "否",
                "张三",
                " 备注保留空格 ",
                None,
                None,
                None,
                None,
            ]
            _append_row(path, values)

            rows = read_watchlist(path)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["仓库"], "openai/gpt-oss")
            self.assertEqual(rows[0]["项目名称"], "中文项目 ")
            self.assertEqual(rows[0]["备注"], " 备注保留空格 ")

    def test_rejects_duplicate_normalized_repository_without_echoing_values(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["SecretOwner/SecretRepo"])
            _append_row(path, ["secretowner/secretrepo"])

            with self.assertRaises(WatchlistError) as caught:
                read_watchlist(path)

            self.assertNotIn("secretowner", str(caught.exception).lower())

    def test_rejects_missing_reordered_unknown_and_duplicate_headers(self):
        mutations = (
            lambda values: values[:-1],
            lambda values: [values[1], values[0]] + values[2:],
            lambda values: values[:5] + ["私密未知表头"] + values[6:],
            lambda values: values[:5] + [values[4]] + values[6:],
        )
        for index, mutate in enumerate(mutations):
            with self.subTest(index=index), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "watchlist.xlsx"
                create_watchlist(path)
                workbook = load_workbook(path)
                worksheet = workbook[SHEET_NAME]
                values = mutate(list(HEADERS))
                for column in range(1, 18):
                    worksheet.cell(1, column).value = (
                        values[column - 1] if column <= len(values) else None
                    )
                workbook.save(path)
                workbook.close()

                with self.assertRaises(WatchlistError) as caught:
                    read_watchlist(path)
                self.assertNotIn("私密未知表头", str(caught.exception))

    def test_rejects_extra_columns_and_nonempty_row_without_repository(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            workbook = load_workbook(path)
            worksheet = workbook[SHEET_NAME]
            worksheet["Q1"] = "extra"
            workbook.save(path)
            workbook.close()
            with self.assertRaises(WatchlistError):
                read_watchlist(path)

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, [None, "orphan secret"])
            with self.assertRaises(WatchlistError) as caught:
                read_watchlist(path)
            self.assertNotIn("orphan secret", str(caught.exception))

    def test_rejects_invalid_repository_formula_and_error_values_safely(self):
        cases = (
            ("not-a-repository", None),
            ("org/tool", "=HYPERLINK(\"https://secret\")"),
            ("org/tool", "#DIV/0!"),
        )
        for repository, suspicious in cases:
            with self.subTest(suspicious=suspicious), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "watchlist.xlsx"
                create_watchlist(path)
                workbook = load_workbook(path)
                worksheet = workbook[SHEET_NAME]
                worksheet["A2"] = repository
                if suspicious is not None:
                    worksheet["L2"] = suspicious
                    if suspicious.startswith("#"):
                        worksheet["L2"].data_type = "e"
                workbook.save(path)
                workbook.close()

                with self.assertRaises(WatchlistError) as caught:
                    read_watchlist(path)
                self.assertNotIn("secret", str(caught.exception).lower())


class WatchlistUpdateTests(unittest.TestCase):
    def test_baseline_updates_version_and_default_date_without_touching_manual_note(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(
                path,
                [
                    "ORG/Tool",
                    "项目",
                    "使用中",
                    "高",
                    "场景",
                    "v1",
                    "2026-07-01",
                    "",
                    "是",
                    "否",
                    "负责人",
                    "人工备注",
                ],
            )

            updates = {"org/tool": {"最新版本": "v2"}}
            original_updates = copy.deepcopy(updates)
            matched = update_automatic_fields(path, updates)

            workbook = load_workbook(path)
            worksheet = workbook[SHEET_NAME]
            self.assertEqual(matched, 1)
            self.assertEqual(worksheet["L2"].value, "人工备注")
            self.assertEqual(worksheet["M2"].value, date.today().isoformat())
            self.assertEqual(worksheet["N2"].value, "v2")
            self.assertEqual(updates, original_updates)
            workbook.close()

    def test_preserves_every_manual_cell_style_formula_comment_and_sheet_controls(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(
                path,
                ["org/tool", "=\"computed\"", "使用中", "高"] + [None] * 12,
            )
            workbook = load_workbook(path)
            worksheet = workbook[SHEET_NAME]
            worksheet["L2"] = "人工"
            worksheet["L2"].comment = Comment("不能丢", "用户")
            worksheet["C2"].font = Font(name="Arial", bold=True, color="00FF0000")
            worksheet["D2"].fill = PatternFill("solid", fgColor="00FFFF00")
            workbook.save(path)
            workbook.close()

            before = load_workbook(path, data_only=False)
            before_sheet = before[SHEET_NAME]
            manual_values = [before_sheet.cell(2, column).value for column in range(1, 13)]
            manual_styles = [
                copy.copy(before_sheet.cell(2, column)._style)
                for column in range(1, 13)
            ]
            validation_ranges = [
                str(item.sqref)
                for item in before_sheet.data_validations.dataValidation
            ]
            before.close()

            update_automatic_fields(
                path,
                {
                    "ORG/TOOL": {
                        "最新版本": "v3",
                        "维护状态": "活跃",
                        "建议动作": "继续关注",
                    }
                },
            )

            after = load_workbook(path, data_only=False)
            after_sheet = after[SHEET_NAME]
            self.assertEqual(
                [after_sheet.cell(2, column).value for column in range(1, 13)],
                manual_values,
            )
            self.assertEqual(
                [after_sheet.cell(2, column)._style for column in range(1, 13)],
                manual_styles,
            )
            self.assertEqual(after_sheet["B2"].data_type, "f")
            self.assertEqual(after_sheet["L2"].comment.text, "不能丢")
            self.assertEqual(after_sheet.freeze_panes, "A2")
            self.assertEqual(after_sheet.auto_filter.ref, "A1:P1")
            self.assertEqual(
                [
                    str(item.sqref)
                    for item in after_sheet.data_validations.dataValidation
                ],
                validation_ranges,
            )
            self.assertEqual(after_sheet["N2"].value, "v3")
            self.assertEqual(after_sheet["O2"].value, "活跃")
            self.assertEqual(after_sheet["P2"].value, "继续关注")
            after.close()

    def test_zip_update_preserves_non_target_entries_rich_text_and_formula_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            workbook = load_workbook(path)
            workbook.create_sheet("占位工作表", 0)["A1"] = "unchanged"
            worksheet = workbook[SHEET_NAME]
            worksheet["A2"] = "org/tool"
            worksheet["B2"] = "=1+1"
            worksheet["C2"] = CellRichText(
                TextBlock(InlineFont(b=True, color="00FF0000"), "富"), "文本"
            )
            worksheet["L2"] = "人工备注"
            worksheet["L2"].comment = Comment("包级保留", "用户")
            workbook.save(path)
            workbook.close()

            sheet_entry = _worksheet_entry(path)
            self.assertNotEqual(sheet_entry, "xl/worksheets/sheet1.xml")
            _inject_formula_cache(path, sheet_entry, "B2", "2")
            before_payloads = _zip_payloads(path)
            before_manual_cells = _worksheet_cells(before_payloads[sheet_entry])

            update_automatic_fields(
                path,
                {
                    "org/tool": {
                        "最新版本": "v2",
                        "维护状态": "活跃",
                        "建议动作": "继续关注",
                    }
                },
            )

            after_payloads = _zip_payloads(path)
            self.assertEqual(set(after_payloads), set(before_payloads))
            for entry_name, before_bytes in before_payloads.items():
                if entry_name != sheet_entry:
                    self.assertEqual(after_payloads[entry_name], before_bytes, entry_name)
            self.assertEqual(
                _worksheet_cells(after_payloads[sheet_entry]), before_manual_cells
            )
            target_root = ET.fromstring(after_payloads[sheet_entry])
            cells = {
                cell.get("r"): cell
                for cell in target_root.findall(".//{{{}}}c".format(_MAIN_NS))
            }
            self.assertEqual(cells["B2"].find("{{{}}}v".format(_MAIN_NS)).text, "2")
            self.assertEqual(
                len(cells["C2"].findall(".//{{{}}}r".format(_MAIN_NS))), 2
            )
            expected = {
                "M2": date.today().isoformat(),
                "N2": "v2",
                "O2": "活跃",
                "P2": "继续关注",
            }
            for reference, text_value in expected.items():
                cell = cells[reference]
                self.assertEqual(cell.get("t"), "inlineStr")
                self.assertIsNone(cell.find("{{{}}}f".format(_MAIN_NS)))
                self.assertIsNone(cell.find("{{{}}}v".format(_MAIN_NS)))
                self.assertEqual(
                    cell.find(".//{{{}}}t".format(_MAIN_NS)).text, text_value
                )

    def test_byte_patch_preserves_markup_and_handles_existing_and_missing_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/tool", "人工内容"])
            sheet_entry = _worksheet_entry(path)
            alternate = (
                b'<mc:AlternateContent><mc:Choice Requires="x14ac">'
                b'<x14ac:futureFeature val="keep"><![CDATA[future <raw>]]>'
                b'</x14ac:futureFeature></mc:Choice>'
                b'<mc:Fallback/></mc:AlternateContent>'
            )
            comment = b"<!--KEEP-COMMENT exact bytes-->"
            instruction = b"<?keep-processing exact-bytes?>"

            def inject_markup(xml_bytes):
                root_start = xml_bytes.index(b"<worksheet")
                root_end = xml_bytes.index(b">", root_start)
                root_open = xml_bytes[root_start : root_end + 1]
                enriched_open = root_open[:-1] + (
                    b' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
                    b' xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"'
                    b' mc:Ignorable="x14ac">'
                )
                xml_bytes = (
                    xml_bytes[:root_start]
                    + enriched_open
                    + xml_bytes[root_end + 1 :]
                )
                row_start = xml_bytes.index(b'<row r="2"')
                row_end = xml_bytes.index(b"</row>", row_start)
                extra_cells = (
                    b'<c r="N2" s="1" ph="1"/>'
                    b'<c r="P2" s="1" t="str"><v>old</v></c>'
                )
                xml_bytes = xml_bytes[:row_end] + extra_cells + xml_bytes[row_end:]
                worksheet_end = xml_bytes.index(b"</worksheet>")
                preserved = comment + instruction + alternate
                return (
                    xml_bytes[:worksheet_end]
                    + preserved
                    + xml_bytes[worksheet_end:]
                )

            _rewrite_zip_payload(path, sheet_entry, inject_markup)
            before = _zip_payloads(path)[sheet_entry]
            root_start = before.index(b"<worksheet")
            root_end = before.index(b">", root_start) + 1
            root_open = before[root_start:root_end]
            row_start = before.index(b'<row r="2"')
            row_end = before.index(b"</row>", row_start) + len(b"</row>")
            before_prefix = before[:row_start]
            before_suffix = before[row_end:]
            version = " <&>\"'\r "

            update_automatic_fields(
                path,
                {
                    "org/tool": {
                        "最新版本": version,
                        "维护状态": "活跃",
                        "建议动作": "继续",
                    }
                },
            )

            with zipfile.ZipFile(path, "r") as archive:
                self.assertIsNone(archive.testzip())
                after = archive.read(sheet_entry)
            self.assertIn(root_open, after)
            after_row_start = after.index(b'<row r="2"')
            after_row_end = after.index(b"</row>", after_row_start) + len(b"</row>")
            row_bytes = after[after_row_start:after_row_end]
            self.assertEqual(after[:after_row_start], before_prefix)
            self.assertEqual(after[after_row_end:], before_suffix)
            self.assertIn(comment, after)
            self.assertIn(instruction, after)
            self.assertIn(alternate, after)
            references = re.findall(br'<c\b[^>]*\br="([A-P]2)"', row_bytes)
            self.assertEqual(references, sorted(references))
            self.assertIn(b'<c r="N2" s="1" ph="1" t="inlineStr">', row_bytes)
            self.assertIn(b'<c r="P2" s="1" t="inlineStr">', row_bytes)
            self.assertIn(
                b"<t xml:space=\"preserve\"> &lt;&amp;&gt;\"'&#13; </t>",
                row_bytes,
            )

            workbook = load_workbook(path, data_only=False)
            worksheet = workbook[SHEET_NAME]
            self.assertEqual(worksheet["N2"].value, version)
            self.assertEqual(worksheet["O2"].value, "活跃")
            self.assertEqual(worksheet["P2"].value, "继续")
            workbook.close()

    def test_rejects_unknown_fields_duplicate_normalized_keys_and_bad_repositories(self):
        cases = (
            {"org/tool": {"备注": "must not write"}},
            {"org/tool": {"未知自动列": "x"}},
            {"ORG/TOOL": {"最新版本": "v1"}, "org/tool": {"最新版本": "v2"}},
            {"not-a-repo": {"最新版本": "v1"}},
        )
        for updates in cases:
            with self.subTest(updates=updates), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "watchlist.xlsx"
                create_watchlist(path)
                _append_row(path, ["org/tool"])
                original = path.read_bytes()
                with self.assertRaises(WatchlistError):
                    update_automatic_fields(path, updates)
                self.assertEqual(path.read_bytes(), original)

    def test_rejects_formula_like_and_unsafe_automatic_values_without_mutating_input(self):
        unsafe_values = (
            "=1+1",
            "+SUM(A1:A2)",
            "-2+3",
            "@cmd",
            ["list"],
            {"nested": "value"},
            float("inf"),
        )
        for value in unsafe_values:
            with self.subTest(value=value), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "watchlist.xlsx"
                create_watchlist(path)
                _append_row(path, ["org/tool"])
                updates = {"ORG/TOOL": {"最新版本": value}}
                original_input = copy.deepcopy(updates)
                original_file = path.read_bytes()

                with self.assertRaises(WatchlistError) as caught:
                    update_automatic_fields(path, updates)

                self.assertNotIn(str(value), str(caught.exception))
                self.assertEqual(updates, original_input)
                self.assertEqual(path.read_bytes(), original_file)

    def test_prevalidates_xml_text_timezone_and_excel_numeric_limits(self):
        invalid_values = (
            "private-marker\x00control",
            "private-marker" + "x" * 32768,
            datetime(2026, 7, 14, 8, tzinfo=timezone.utc),
            2**53,
            float("nan"),
            5e-324,
        )
        for value in invalid_values:
            with self.subTest(value_type=type(value).__name__), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "watchlist.xlsx"
                create_watchlist(path)
                _append_row(path, ["org/tool"])
                original = path.read_bytes()
                with patch("watchlist.exclusive_file_lock") as lock:
                    with self.assertRaises(WatchlistError) as caught:
                        update_automatic_fields(
                            path, {"org/tool": {"最新版本": value}}
                        )
                lock.assert_not_called()
                self.assertNotIn("private-marker", str(caught.exception))
                self.assertEqual(path.read_bytes(), original)

    def test_unknown_repository_is_ignored_with_deterministic_zero_result(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/known"])
            original = path.read_bytes()

            matched = update_automatic_fields(
                path, {"org/unknown": {"最新版本": "v9"}}
            )

            self.assertEqual(matched, 0)
            self.assertEqual(path.read_bytes(), original)

    def test_accepts_safe_date_scalar_in_automatic_field(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/tool"])

            matched = update_automatic_fields(
                path,
                {"org/tool": {"最近检查时间": date(2026, 7, 14)}},
            )

            self.assertEqual(matched, 1)
            self.assertEqual(
                read_watchlist(path)[0]["最近检查时间"],
                "2026-07-14",
            )

    def test_bad_header_or_duplicate_sheet_repository_aborts_without_change(self):
        for mutation in ("header", "duplicate"):
            with self.subTest(mutation=mutation), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "watchlist.xlsx"
                create_watchlist(path)
                _append_row(path, ["org/tool"])
                workbook = load_workbook(path)
                worksheet = workbook[SHEET_NAME]
                if mutation == "header":
                    worksheet["P1"] = "bad"
                else:
                    worksheet.append(["ORG/TOOL"])
                workbook.save(path)
                workbook.close()
                original = path.read_bytes()

                with self.assertRaises(WatchlistError):
                    update_automatic_fields(
                        path, {"org/tool": {"最新版本": "v2"}}
                    )

                self.assertEqual(path.read_bytes(), original)

    def test_atomic_replace_failure_preserves_original_and_cleans_temp_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/tool"])
            original = path.read_bytes()

            with patch("storage.os.replace", side_effect=OSError("replace failed")):
                with self.assertRaises(WatchlistError):
                    update_automatic_fields(
                        path, {"org/tool": {"最新版本": "v2"}}
                    )

            self.assertEqual(path.read_bytes(), original)
            self.assertEqual(list(path.parent.glob("*.tmp.xlsx")), [])

    def test_concurrent_updates_do_not_lose_each_others_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/one"])
            _append_row(path, ["org/two"])
            context = multiprocessing.get_context("fork")
            start_event = context.Event()
            processes = [
                context.Process(
                    target=_concurrent_update,
                    args=(str(path), start_event, "org/one", "v1"),
                ),
                context.Process(
                    target=_concurrent_update,
                    args=(str(path), start_event, "org/two", "v2"),
                ),
            ]
            for process in processes:
                process.start()
            start_event.set()
            for process in processes:
                process.join(10)
                self.assertEqual(process.exitcode, 0)

            rows = {row["仓库"]: row for row in read_watchlist(path)}
            self.assertEqual(rows["org/one"]["最新版本"], "v1")
            self.assertEqual(rows["org/two"]["最新版本"], "v2")


class WatchlistStorageSafetyTests(unittest.TestCase):
    def test_update_preserves_existing_mode_and_new_template_is_private(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            self.assertEqual(stat.S_IMODE(path.stat().st_mode), 0o600)
            _append_row(path, ["org/tool"])
            path.chmod(0o644)

            update_automatic_fields(
                path, {"org/tool": {"最新版本": "v2"}}
            )

            self.assertEqual(stat.S_IMODE(path.stat().st_mode), 0o644)

    def test_explicit_unlock_failure_after_replace_is_not_a_transaction_failure(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/tool"])
            real_flock = fcntl.flock

            def fail_only_explicit_unlock(file_descriptor, operation):
                if operation == fcntl.LOCK_UN:
                    raise OSError("explicit unlock failed")
                return real_flock(file_descriptor, operation)

            with patch("storage.fcntl.flock", side_effect=fail_only_explicit_unlock):
                matched = update_automatic_fields(
                    path, {"org/tool": {"最新版本": "v2"}}
                )

            self.assertEqual(matched, 1)
            self.assertEqual(read_watchlist(path)[0]["最新版本"], "v2")

    def test_target_symlinks_are_rejected_by_all_public_operations(self):
        operations = (
            lambda path: create_watchlist(path),
            lambda path: read_watchlist(path),
            lambda path: update_automatic_fields(
                path, {"org/tool": {"最新版本": "v2"}}
            ),
        )
        for operation in operations:
            with self.subTest(operation=operation), tempfile.TemporaryDirectory() as tmp:
                real_path = Path(tmp) / "real.xlsx"
                link_path = Path(tmp) / "alias.xlsx"
                create_watchlist(real_path)
                _append_row(real_path, ["org/tool"])
                original = real_path.read_bytes()
                link_path.symlink_to(real_path)

                with self.assertRaises(WatchlistError):
                    operation(link_path)

                self.assertTrue(link_path.is_symlink())
                self.assertEqual(real_path.read_bytes(), original)

    def test_lock_symlink_is_rejected_without_touching_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            create_watchlist(path)
            _append_row(path, ["org/tool"])
            original = path.read_bytes()
            sentinel = Path(tmp) / "sentinel"
            sentinel.write_text("do not touch", encoding="utf-8")
            lock_path = lock_path_for(path)
            lock_path.unlink()
            lock_path.symlink_to(sentinel)

            with self.assertRaises(WatchlistError):
                update_automatic_fields(
                    path, {"org/tool": {"最新版本": "v2"}}
                )

            self.assertEqual(path.read_bytes(), original)
            self.assertEqual(sentinel.read_text(encoding="utf-8"), "do not touch")

    def test_create_race_uses_no_clobber_install_and_keeps_race_winner(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "watchlist.xlsx"
            competitor = b"race-winner"
            real_link = os.link

            def install_competitor_then_link(source, target):
                Path(target).write_bytes(competitor)
                return real_link(source, target)

            with patch("storage.os.link", side_effect=install_competitor_then_link):
                with self.assertRaises(WatchlistError):
                    create_watchlist(path)

            self.assertEqual(path.read_bytes(), competitor)
            self.assertEqual(list(path.parent.glob("*.tmp.xlsx")), [])


if __name__ == "__main__":
    unittest.main()
