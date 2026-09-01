#!/usr/bin/env python3
"""Offline contract tests for the GitHub trend radar command line."""

import contextlib
import copy
import io
import json
import sys
import tempfile
import threading
import unittest
from contextlib import contextmanager
from datetime import date, datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

from config import RadarPaths
from fetch_github import parse_search_rows
from fetch_ossinsight import parse_trend_rows
from fetch_trending import parse_trending_html
from http_client import SourceError
from models import RepositoryRecord
from radar import (
    RadarApp, RadarServices, _has_chinese, _problem_solved,
    _repository_purpose, _reusable_content, _reusability,
    cleanup_stale_latest_reports,
    build_parser, main, should_save,
)
from storage import save_complete_report
from watchlist import HEADERS, SHEET_NAME, create_watchlist


FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


def load_json(name):
    return json.loads((FIXTURES_DIR / name).read_text(encoding="utf-8"))


def records_fixture():
    top = parse_search_rows(load_json("github_search.json"))
    day = parse_trend_rows(load_json("ossinsight_24h.json"), "24h")
    week = parse_trend_rows(load_json("ossinsight_7d.json"), "7d")
    trending = parse_trending_html(
        (FIXTURES_DIR / "github_trending.html").read_text(encoding="utf-8"),
        "weekly",
    )
    return top, day, week, trending


def detail_for(full_name):
    for record in records_fixture()[0]:
        if record.full_name == full_name:
            result = copy.deepcopy(record)
            result.latest_release = "v1.2.3" if full_name == "org/tool" else ""
            result.latest_release_at = "2026-07-13T00:00:00Z"
            result.source_records = [
                {"source": "github", "scope": "repository_detail"}
            ]
            return result
    return RepositoryRecord(
        full_name=full_name,
        total_stars=100,
        license="MIT",
        description="AI coding agent command line tool",
        topics=["ai", "cli", "ai-agent"],
        pushed_at="2026-07-13T00:00:00Z",
        source_records=[{"source": "github", "scope": "repository_detail"}],
    )


def good_services(**overrides):
    top_records, day, week, trending = records_fixture()
    values = dict(
        fetch_top=lambda top=20: copy.deepcopy(top_records[:top]),
        fetch_details=lambda name: detail_for(name),
        fetch_star_count=lambda name: RepositoryRecord(
            full_name=name,
            total_stars=100,
            source_records=[{
                "source": "github_web",
                "scope": "repository_star_fallback",
            }],
        ),
        fetch_readme=lambda name: {},
        fetch_trend=lambda period: copy.deepcopy(day if period == "24h" else week),
        fetch_trending=lambda period="weekly": copy.deepcopy(trending),
    )
    values.update(overrides)
    return RadarServices(**values)


def report_args(category=None, save=False, top=20):
    return SimpleNamespace(command="report", category=category, save=save, top=top)


def append_watchlist(path, full_name, check="是"):
    workbook = load_workbook(path)
    worksheet = workbook[SHEET_NAME]
    values = [None] * len(HEADERS)
    values[0] = full_name
    values[1] = full_name.rsplit("/", 1)[-1]
    values[8] = check
    worksheet.append(values)
    workbook.save(path)
    workbook.close()


class ParserTests(unittest.TestCase):
    def test_default_workspace_uses_planned_chinese_output_root(self):
        app = RadarApp(services=good_services())

        self.assertEqual(
            app.workspace,
            SCRIPTS_DIR.parents[3]
            / "GitHub开源趋势雷达",
        )

    def test_explicit_workspace_is_the_output_root_without_extra_nesting(self):
        with tempfile.TemporaryDirectory() as tmp:
            app = RadarApp(workspace=tmp, services=good_services())

            self.assertEqual(app._paths().root, Path(tmp))
            self.assertEqual(
                app._paths().latest_data,
                Path(tmp)
                / "最新报告"
                / "原始数据-{}.json".format(app._paths().run_date),
            )

    def test_output_root_uses_the_direct_asset_directory(self):
        with tempfile.TemporaryDirectory() as tmp, patch("radar.RadarApp") as app_type:
            app_type.return_value.run_watchlist.return_value = 0

            self.assertEqual(main(["--output-root", tmp, "watchlist"]), 0)

            app_type.assert_called_once_with(workspace=Path(tmp))
            app_type.return_value.run_watchlist.assert_called_once_with()

    def test_report_defaults_and_save_policy(self):
        parser = build_parser()
        full = parser.parse_args(["report"])
        preview = parser.parse_args(["report", "--category", "AI"])
        saved_category = parser.parse_args(
            ["report", "--category", "AI", "--save", "--top", "7"]
        )
        self.assertEqual(full.top, 20)
        self.assertTrue(should_save(full))
        self.assertFalse(should_save(preview))
        self.assertTrue(should_save(saved_category))
        self.assertEqual(saved_category.top, 7)

    def test_report_category_accepts_only_fixed_primary_categories(self):
        parser = build_parser()
        self.assertEqual(parser.parse_args(["report", "--category", "AI"]).category, "AI")
        for value in ("ai", "自定义分类"):
            with self.subTest(value=value):
                stderr = io.StringIO()
                with contextlib.redirect_stderr(stderr), self.assertRaises(SystemExit):
                    parser.parse_args(["report", "--category", value])

    def test_parser_supports_exact_four_subcommands(self):
        parser = build_parser()
        self.assertEqual(parser.parse_args(["repo", "OpenAI/Codex"]).command, "repo")
        self.assertEqual(
            parser.parse_args(["compare", "org/old", "org/new"]).command,
            "compare",
        )
        self.assertEqual(parser.parse_args(["watchlist"]).command, "watchlist")

    def test_global_workspace_is_project_root_and_dispatches_to_fixed_output(self):
        with tempfile.TemporaryDirectory() as tmp, patch("radar.RadarApp") as app_type:
            app_type.return_value.run_watchlist.return_value = 9

            self.assertEqual(main(["--workspace", tmp, "watchlist"]), 9)

            app_type.assert_called_once_with(
                workspace=Path(tmp)
                / "GitHub开源趋势雷达"
            )
            app_type.return_value.run_watchlist.assert_called_once_with()

    def test_parser_rejects_unsafe_values_without_echoing_them(self):
        cases = (
            ["report", "--top", "0"],
            ["report", "--top", "101"],
            ["report", "--category", "bad\ncategory"],
            ["report", "--category", "x" * 81],
            ["--workspace", "bad\nworkspace", "watchlist"],
            ["repo", "https://github.com/org/repo?token=secret"],
            ["repo", "owner/.."],
            ["compare", "org/ok", "bad/repo/extra"],
        )
        for argv in cases:
            with self.subTest(argv=argv):
                stderr = io.StringIO()
                with contextlib.redirect_stderr(stderr), self.assertRaises(SystemExit):
                    build_parser().parse_args(argv)
                self.assertNotIn("secret", stderr.getvalue())
                self.assertNotIn("bad/repo/extra", stderr.getvalue())

    def test_should_save_is_false_for_non_report_and_main_dispatches_once(self):
        self.assertFalse(should_save(SimpleNamespace(
            command="repo", category=None, save=True
        )))
        with patch("radar.run_watchlist", return_value=7) as selected, patch(
            "radar.run_report"
        ) as report, patch("radar.run_repo") as repo, patch(
            "radar.run_compare"
        ) as compare:
            self.assertEqual(main(["watchlist"]), 7)
        selected.assert_called_once_with()
        report.assert_not_called()
        repo.assert_not_called()
        compare.assert_not_called()

    def test_main_returns_nonzero_for_handled_invalid_input(self):
        stderr = io.StringIO()
        with contextlib.redirect_stderr(stderr):
            code = main(["report", "--top", "0"])
        self.assertNotEqual(code, 0)


class PointCommandTests(unittest.TestCase):
    def app(self, workspace, services):
        return RadarApp(
            workspace=workspace,
            services=services,
            today=lambda: date(2026, 7, 14),
            now=lambda: datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
        )

    def capture(self, function, *args):
        stdout, stderr = io.StringIO(), io.StringIO()
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            code = function(*args)
        return code, stdout.getvalue(), stderr.getvalue()

    def test_repo_outputs_strict_json_with_missing_signal_confidence_and_no_writes(self):
        with tempfile.TemporaryDirectory() as tmp:
            code, stdout, stderr = self.capture(
                self.app(tmp, good_services()).run_repo, "org/tool"
            )
            self.assertEqual(code, 0, stderr)
            result = json.loads(stdout)
            self.assertEqual(result["repository"]["full_name"], "org/tool")
            self.assertTrue(result["reusability"]["missing_signals"])
            self.assertNotEqual(result["reusability"]["confidence"], "高")
            self.assertFalse(any(Path(tmp).iterdir()))

    def test_repo_source_error_is_safe_nonzero_and_does_not_emit_json(self):
        def failed(name):
            raise SourceError("Authorization Bearer top-secret")

        with tempfile.TemporaryDirectory() as tmp:
            code, stdout, stderr = self.capture(
                self.app(tmp, good_services(fetch_details=failed)).run_repo,
                "org/tool",
            )
            self.assertNotEqual(code, 0)
            self.assertEqual(stdout, "")
            self.assertNotIn("top-secret", stderr)
            self.assertFalse(any(Path(tmp).iterdir()))

    def test_repo_does_not_modify_injected_record(self):
        original = detail_for("org/tool")
        snapshot = copy.deepcopy(original)
        with tempfile.TemporaryDirectory() as tmp:
            code, _, stderr = self.capture(
                self.app(
                    tmp, good_services(fetch_details=lambda name: original)
                ).run_repo,
                "org/tool",
            )
        self.assertEqual(code, 0, stderr)
        self.assertEqual(original, snapshot)

    def test_repo_distinguishes_known_zero_from_missing_detail_signals(self):
        cases = (
            ("available", True, {"releases": 1.0, "community": 0.0}, []),
            ("none_or_unavailable", True, {"releases": 0.0, "community": 0.0}, []),
            ("release_fetch_failed", True, {"community": 0.0}, ["releases"]),
            (None, True, {"releases": 0.0, "community": 0.0}, []),
            (None, False, {}, ["community", "releases"]),
        )
        for release_status, detail_collected, expected_signals, expected_missing in cases:
            with self.subTest(
                status=release_status, detail_collected=detail_collected
            ), tempfile.TemporaryDirectory() as tmp:
                source_records = []
                if detail_collected:
                    source_records = [{
                        "source": "github",
                        "scope": "repository_detail",
                    }]
                    if release_status is not None:
                        source_records[0]["release"] = release_status
                record = RepositoryRecord(
                    full_name="org/zero",
                    total_stars=0,
                    forks=0,
                    license="MIT",
                    source_records=source_records,
                )
                code, stdout, stderr = self.capture(
                    self.app(
                        tmp, good_services(fetch_details=lambda name: record)
                    ).run_repo,
                    "org/zero",
                )
                self.assertEqual(code, 0, stderr)
                reuse = json.loads(stdout)["reusability"]
                for name, value in expected_signals.items():
                    self.assertEqual(reuse["signals"][name], value)
                for name in expected_missing:
                    self.assertIn(name, reuse["missing_signals"])
                if not detail_collected:
                    self.assertNotIn("community", reuse["signals"])

    def test_compare_outputs_scores_overlap_and_evidence_gated_relation(self):
        with tempfile.TemporaryDirectory() as tmp:
            code, stdout, stderr = self.capture(
                self.app(tmp, good_services()).run_compare, "org/old", "org/new"
            )
            self.assertEqual(code, 0, stderr)
            result = json.loads(stdout)
            self.assertIn("reusability", result["old"])
            self.assertIn("ai agent", result["tag_overlap"])
            self.assertIn("cli", result["tag_overlap"])
            self.assertNotEqual(result["replacement"]["relation"], "直接替代")
            self.assertEqual(result["replacement"]["evidence"], [])
            self.assertTrue(result["evidence_gaps"])
            self.assertFalse(any(Path(tmp).iterdir()))

    def test_compare_single_fetch_failure_is_safe_and_does_not_persist(self):
        def detail(name):
            if name == "org/new":
                raise SourceError("token=secret")
            return detail_for(name)

        with tempfile.TemporaryDirectory() as tmp:
            code, stdout, stderr = self.capture(
                self.app(tmp, good_services(fetch_details=detail)).run_compare,
                "org/old",
                "org/new",
            )
            self.assertNotEqual(code, 0)
            self.assertEqual(stdout, "")
            self.assertNotIn("token=secret", stderr)
            self.assertFalse(any(Path(tmp).iterdir()))

    def test_watchlist_updates_once_after_all_attempts_and_isolates_failures(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/tool")
            append_watchlist(paths.watchlist_file, "org/fail")
            append_watchlist(paths.watchlist_file, "org/skipped", check="否")
            workbook = load_workbook(paths.watchlist_file)
            original_manual = [
                row[:12]
                for row in workbook[SHEET_NAME].iter_rows(values_only=True)
            ]
            workbook.close()
            attempts, update_calls = [], []

            def detail(name):
                attempts.append(name)
                if name == "org/fail":
                    raise SourceError("Bearer hidden")
                return detail_for(name)

            def update(path, values):
                update_calls.append(copy.deepcopy(values))
                from watchlist import update_automatic_fields
                return update_automatic_fields(path, values)

            services = good_services(
                fetch_details=detail, update_watchlist=update
            )
            code, stdout, stderr = self.capture(
                self.app(tmp, services).run_watchlist
            )
            result = json.loads(stdout)
            self.assertNotEqual(code, 0)
            self.assertEqual(attempts, ["org/tool", "org/fail"])
            self.assertEqual(len(update_calls), 1)
            self.assertEqual(result["updated"], 1)
            self.assertEqual(result["status"], "partial")
            self.assertEqual(result["failed_repositories"], ["org/fail"])
            self.assertNotIn("Bearer", stdout + stderr)
            workbook = load_workbook(paths.watchlist_file)
            worksheet = workbook[SHEET_NAME]
            current_manual = [
                row[:12] for row in worksheet.iter_rows(values_only=True)
            ]
            self.assertEqual(current_manual, original_manual)
            self.assertEqual(worksheet["N2"].value, "v1.2.3")
            workbook.close()

    def test_watchlist_all_target_failures_are_degraded(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/fail")
            services = good_services(fetch_details=lambda name: (_ for _ in ()).throw(
                SourceError("Bearer hidden")
            ))

            code, stdout, stderr = self.capture(
                self.app(tmp, services).run_watchlist
            )
            result = json.loads(stdout)

            self.assertNotEqual(code, 0)
            self.assertEqual(result["status"], "degraded")
            self.assertEqual(result["updated"], 0)
            self.assertNotIn("Bearer", stdout + stderr)

    def test_repo_and_compare_reject_nonfinite_injected_numeric_output(self):
        invalid_records = {
            "org/nan": RepositoryRecord(
                full_name="org/nan", total_stars=float("nan")
            ),
            "org/inf": RepositoryRecord(
                full_name="org/inf", forks=float("inf")
            ),
        }

        def detail(name):
            return invalid_records.get(name, detail_for(name))

        with tempfile.TemporaryDirectory() as tmp:
            app = self.app(tmp, good_services(
                fetch_details=detail,
                fetch_trend=lambda period: [],
            ))
            code, stdout, stderr = self.capture(app.run_repo, "org/nan")
            self.assertNotEqual(code, 0)
            self.assertEqual(stdout, "")
            self.assertNotIn("NaN", stderr)

            code, stdout, stderr = self.capture(
                app.run_compare, "org/tool", "org/inf"
            )
            self.assertNotEqual(code, 0)
            self.assertEqual(stdout, "")
            self.assertNotIn("Infinity", stderr)

    def test_watchlist_unexpected_failure_does_not_update(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/tool")
            before = paths.watchlist_file.read_bytes()
            update_calls = []

            def crashed(name):
                raise RuntimeError("token=unexpected")

            services = good_services(
                fetch_details=crashed,
                update_watchlist=lambda *args: update_calls.append(args),
            )
            code, stdout, stderr = self.capture(
                self.app(tmp, services).run_watchlist
            )
            self.assertNotEqual(code, 0)
            self.assertEqual(update_calls, [])
            self.assertEqual(paths.watchlist_file.read_bytes(), before)
            self.assertNotIn("unexpected", stdout + stderr)


class ReportPreviewTests(unittest.TestCase):
    def test_english_repository_description_never_leaks_into_core_copy(self):
        record = RepositoryRecord(
            full_name="org/english",
            description="An English-only agent platform for developers.",
            primary_category="AI",
            repository_type="可运行软件",
        )

        purpose = _repository_purpose(record)
        problem = _problem_solved(record)

        self.assertTrue(_has_chinese(purpose))
        self.assertTrue(_has_chinese(problem))
        self.assertNotIn("English-only", purpose + problem)

    def test_reusable_content_is_rendered_as_chinese_summary(self):
        record = RepositoryRecord(
            full_name="org/agent",
            primary_category="AI",
            repository_type="SDK 或库",
            secondary_tags=["AI Agent", "MCP"],
            description="An English repository description.",
        )
        self.assertEqual(
            _reusable_content(record),
            "AI领域的SDK 或库，可用于AI Agent、MCP相关能力的评估与集成。",
        )

    def test_pinned_reusable_project_is_kept_when_details_are_missing(self):
        records = [
            RepositoryRecord(
                full_name="org/high-{}".format(index),
                repository_type="可运行软件",
                license="MIT",
                pushed_at="2026-07-13T00:00:00Z",
                source_records=[{"source": "github", "scope": "repository_detail"}],
            )
            for index in range(3)
        ]
        pinned = RepositoryRecord(
            full_name="org/pinned",
            repository_type="可运行软件",
        )
        app = RadarApp(services=good_services())

        _, reusable = app._category_and_reuse(
            records + [pinned],
            3,
            datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
            ["org/pinned"],
        )

        self.assertIn("org/pinned", {row["repo"] for row in reusable})
        pinned_row = next(row for row in reusable if row["repo"] == "org/pinned")
        self.assertIn("固定关注项目", "；".join(pinned_row["risks"]))
        self.assertIn("purpose", pinned_row)
        self.assertIn("problem_solved", pinned_row)
        self.assertIn("total_stars", pinned_row)

    def test_english_readme_falls_back_to_chinese_summary(self):
        record = RepositoryRecord(
            full_name="org/english",
            primary_category="AI",
            repository_type="可运行软件",
            description="English project",
        )
        services = good_services(fetch_readme=lambda name: {
            "features": ["Zero configuration desktop agent"],
            "use_cases": ["Automate local workflows"],
            "source": ["README:main/README.md"],
        })
        app = RadarApp(services=services)

        _, reusable = app._category_and_reuse(
            [record], 1,
            datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
        )

        row = reusable[0]
        self.assertTrue(_has_chinese(row["purpose"]))
        self.assertTrue(_has_chinese(row["problem_solved"]))
        self.assertIn("回退中文分类摘要", row["description_source"])

    def test_reusability_uses_new_signal_weights_and_exposes_coverage(self):
        record = RepositoryRecord(
            full_name="org/complete",
            description="x" * 200,
            repository_type="SDK 或库",
            total_stars=100,
            forks=10,
            primary_language="Python",
            topics=["ci", "sdk", "library", "python", "framework"],
            license="MIT",
            pushed_at="2026-07-13T00:00:00Z",
            latest_release="v1.2.3",
            source_records=[
                {"source": "github", "scope": "repository_detail"},
                {"source": "github_latest_release", "status": "available"},
            ],
        )

        result = _reusability(
            record,
            datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
        )

        self.assertEqual(result["score"], 100)
        self.assertEqual(result["components"]["docs"], 15.0)
        self.assertEqual(result["components"]["integration"], 15.0)
        self.assertEqual(result["components"]["ci"], 5.0)
        self.assertEqual(result["missing_signals"], [])
        self.assertEqual(result["confidence"], "高")

    def test_model_returns_new_featured_and_market_sections(self):
        def record(name, category, stars):
            return RepositoryRecord(
                full_name=name,
                primary_category=category,
                repository_type="可运行软件",
                total_stars=stars,
                source_records=[{"source": "github", "scope": "search"}],
            )

        popular = record("org/popular", "其他", 500000)
        ai = record("org/ai", "AI", 5000)
        dev = record("org/dev", "开发工具", 4000)
        rankings = {
            "total": [],
            "daily": [
                {"repo": "org/dev", "value": 3, "source": "OSSInsight 24h external"}
            ],
            "weekly": [
                {
                    "repo": "org/ai", "value": 1200,
                    "source": "local history 7d", "primary_category": "AI",
                },
                {
                    "repo": "org/dev", "value": 700,
                    "source": "local history 7d", "primary_category": "开发工具",
                },
                {
                    "repo": "org/other", "value": 400,
                    "source": "local history 7d", "primary_category": "其他",
                },
            ],
            "acceleration": [],
        }
        reusable = [
            {
                "repo": "org/ai", "repository_type": "可运行软件",
                "runnable": True, "purpose": "AI 项目。",
                "problem_solved": "解决 AI 应用问题。", "description_source": "分类摘要",
                "total_stars": 5000, "reuse": "源码", "integration": "源码评估",
                "score": 1, "license": "MIT", "maintenance": "活跃",
                "risks": [], "actions": [],
            }
        ]
        app = RadarApp(services=good_services())
        model = app._model(
            [popular, ai, dev], [], [], [], [], report_args(),
            datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
            objective_parts=(
                [popular, ai, dev], rankings, [popular, ai, dev],
                [{"category": "AI", "facts": ["本期纳入 1 个仓库。"], "change": None, "confidence": "中"}],
                reusable,
            ),
        )

        self.assertIsInstance(model, dict)
        self.assertEqual([row["repo"] for row in model["featured_projects"]], [
            "org/ai", "org/dev", "org/popular",
        ])
        self.assertTrue(any(
            "AI 方向本周增长集中" in item
            for item in model["overview"]["inferences"]
        ))
        self.assertTrue(model["market_trends"])
        self.assertIn("1200 个 Star", model["market_trends"][0]["conclusion"])
        self.assertEqual(model["market_trends"][0]["evidence"][0]["source"], "local history 7d")

    def app(self, workspace, services):
        return RadarApp(
            workspace=workspace,
            services=services,
            today=lambda: date(2026, 7, 14),
            now=lambda: datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
        )

    def capture_report(self, app, args):
        stdout, stderr = io.StringIO(), io.StringIO()
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            code = app.run_report(args)
        return code, stdout.getvalue(), stderr.getvalue()

    def test_category_preview_initializes_only_config_and_never_persists_run_outputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            code, stdout, stderr = self.capture_report(
                self.app(tmp, good_services()), report_args(category="AI")
            )
            paths = RadarPaths(Path(tmp), "2026-07-14")
            self.assertEqual(code, 0, stderr)
            self.assertIn("# GitHub 趋势雷达日报", stdout)
            self.assertIn("暂无本地历史", stdout)
            self.assertTrue(paths.categories_file.exists())
            self.assertFalse(paths.watchlist_file.exists())
            self.assertFalse(paths.archive_report.exists())
            self.assertFalse(paths.latest_report.exists())
            self.assertFalse(paths.history_file.exists())

    def test_history_availability_fact_matches_model_and_rendered_report(self):
        history = [{
            "date": "2026-07-13",
            "at": "2026-07-13T08:30:00+00:00",
            "repo": "org/tool",
            "stars": 1900,
            "primary_category": "AI",
        }]
        with tempfile.TemporaryDirectory() as empty_tmp:
            empty = self.app(
                empty_tmp, good_services()
            )._collect_report(report_args(category="AI"))
            self.assertTrue(any(
                "暂无本地历史" in fact
                and "不计算" in fact
                for fact in empty[1]["overview"]["facts"]
            ))
            self.assertIn("暂无本地历史", empty[2])

        with tempfile.TemporaryDirectory() as history_tmp:
            (Path(history_tmp) / "运行状态" / "历史指标").mkdir(parents=True)
            populated = self.app(
                history_tmp,
                good_services(load_history=lambda root: copy.deepcopy(history)),
            )._collect_report(report_args(category="AI"))
            self.assertFalse(any(
                "暂无本地历史" in fact
                for fact in populated[1]["overview"]["facts"]
            ))
            self.assertNotIn("暂无本地历史", populated[2])

    def test_period_status_is_independent_and_total_created_growth_are_isolated(self):
        def period(period):
            if period == "24h":
                raise SourceError("Bearer hidden")
            return good_services().fetch_trend(period)

        with tempfile.TemporaryDirectory() as tmp:
            collected = self.app(
                tmp, good_services(fetch_trend=period)
            )._collect_report(report_args())
            self.assertIsNotNone(collected)
            model = collected[1]
            statuses = {
                row["source"]: row["status"]
                for row in model["metadata"]["source_status"]
            }
            self.assertEqual(statuses["ossinsight_24h"], "degraded")
            self.assertEqual(statuses["ossinsight_7d"], "ok")
            self.assertEqual(model["rankings"]["total"][0]["value"], 2000)
            self.assertEqual(model["rankings"]["weekly"][0]["value"], 640)
            self.assertEqual(model["rankings"]["daily"], [])
            self.assertNotEqual(
                model["rankings"]["weekly"][0]["value"],
                "2026-01-01T00:00:00Z",
            )

    def test_ranking_rows_keep_metric_meaning_and_repository_explanation(self):
        with tempfile.TemporaryDirectory() as tmp:
            collected = self.app(tmp, good_services())._collect_report(report_args())
            model = collected[1]
            total = model["rankings"]["total"][0]
            daily = model["rankings"]["daily"][0]
            weekly = model["rankings"]["weekly"][0]

            for row in (total, daily, weekly):
                self.assertEqual(row["primary_category"], "AI")
                self.assertEqual(row["repository_type"], "可运行软件")
                self.assertTrue(_has_chinese(row["purpose"]))
                self.assertIn("AI 应用", row["problem_solved"])
                self.assertNotIn("AI coding CLI", row["purpose"] + row["problem_solved"])
                self.assertTrue(row["data_scope"])

            self.assertEqual(total["metric_name"], "当前累计 Star")
            self.assertEqual(daily["metric_name"], "过去24小时新增 Star")
            self.assertEqual(daily["total_stars"], 2000)
            self.assertEqual(weekly["metric_name"], "过去7天新增 Star")
            self.assertEqual(weekly["total_stars"], 2000)
            self.assertEqual(total["unit"], "个")
            self.assertEqual(daily["unit"], "个")

    def test_trending_selected_set_is_preserved_but_not_used_as_precise_growth(self):
        def no_oss(period):
            raise SourceError("unavailable")

        with tempfile.TemporaryDirectory() as tmp:
            collected = self.app(
                tmp, good_services(fetch_trend=no_oss)
            )._collect_report(report_args())
            model = collected[1]
            self.assertEqual(model["rankings"]["daily"], [])
            self.assertEqual(model["rankings"]["weekly"], [])
            tool = next(
                row for row in model["repositories"]
                if row["full_name"] == "org/tool"
            )
            self.assertEqual(tool["stars_7d_external"], 640)
            self.assertTrue(any(
                row.get("source") == "github_trending"
                and row.get("scope") == "selected_set"
                for row in tool["source_records"]
            ))

    def test_trending_only_repo_is_not_ranked_when_oss_period_succeeds_for_others(self):
        trending_only = RepositoryRecord(
            full_name="other/selected",
            total_stars=999,
            stars_7d_external=999,
            source_records=[{
                "source": "github_trending",
                "scope": "selected_set",
                "period": "weekly",
            }],
        )
        with tempfile.TemporaryDirectory() as tmp:
            services = good_services(fetch_trending=lambda period="weekly": [
                copy.deepcopy(trending_only)
            ])
            collected = self.app(tmp, services)._collect_report(report_args())
            weekly_names = {
                row["repo"] for row in collected[1]["rankings"]["weekly"]
            }
            total_names = {
                row["repo"] for row in collected[1]["rankings"]["total"]
            }
            self.assertIn("org/tool", weekly_names)
            self.assertNotIn("other/selected", weekly_names)
            self.assertNotIn("other/selected", total_names)

    def test_each_discovery_source_error_isolated_and_sensitive_text_is_absent(self):
        names = ("github", "oss24", "oss7", "trending")
        for failed_name in names:
            with self.subTest(source=failed_name), tempfile.TemporaryDirectory() as tmp:
                def top(top=20):
                    if failed_name == "github":
                        raise SourceError("token=never-print")
                    return good_services().fetch_top(top=top)

                def period(period):
                    if (failed_name, period) in {("oss24", "24h"), ("oss7", "7d")}:
                        raise SourceError("Authorization Bearer hidden")
                    return good_services().fetch_trend(period)

                def trending(period="weekly"):
                    if failed_name == "trending":
                        raise SourceError("https://secret.invalid/?token=hidden")
                    return good_services().fetch_trending(period=period)

                code, stdout, stderr = self.capture_report(
                    self.app(tmp, good_services(
                        fetch_top=top,
                        fetch_trend=period,
                        fetch_trending=trending,
                    )),
                    report_args(category="AI"),
                )
                self.assertEqual(code, 0, stderr)
                self.assertNotIn("never-print", stdout + stderr)
                self.assertNotIn("Bearer", stdout + stderr)
                self.assertNotIn("secret.invalid", stdout + stderr)
                if failed_name != "github":
                    self.assertNotIn("警告：", stdout)
                    self.assertIn("警告：", stderr)

    def test_detail_failures_are_partial_and_do_not_discard_search_records(self):
        def detail(name):
            raise SourceError("token=detail-secret")

        with tempfile.TemporaryDirectory() as tmp:
            collected = self.app(
                tmp, good_services(fetch_details=detail)
            )._collect_report(report_args())
            model = collected[1]
            statuses = {
                row["source"]: row["status"]
                for row in model["metadata"]["source_status"]
            }
            self.assertEqual(statuses["github_details"], "degraded")
            self.assertTrue(model["rankings"]["total"])
            rendered = collected[2]
            self.assertNotIn("detail-secret", rendered)

    def test_detail_fetch_order_prioritizes_growth_rankings(self):
        github = [RepositoryRecord(full_name="org/total", total_stars=1000)]
        day = [
            RepositoryRecord(full_name="org/day-low", stars_24h_external=2),
            RepositoryRecord(full_name="org/day-high", stars_24h_external=9),
        ]
        week = [
            RepositoryRecord(full_name="org/week-high", stars_7d_external=20),
            RepositoryRecord(full_name="org/day-high", stars_7d_external=10),
        ]
        trending = [RepositoryRecord(full_name="org/trending", total_stars=500)]

        names = RadarApp._detail_names(github, day, week, trending, top=2)

        self.assertEqual(names[:5], [
            "org/day-high", "org/day-low", "org/week-high",
            "org/trending", "org/total",
        ])

    def test_detail_failure_uses_public_star_fallback_for_growth_rows(self):
        def failed_detail(name):
            raise SourceError("rate limited", status_code=403)

        with tempfile.TemporaryDirectory() as tmp:
            collected = self.app(
                tmp, good_services(fetch_details=failed_detail)
            )._collect_report(report_args())
            model = collected[1]
            statuses = {
                row["source"]: row["status"]
                for row in model["metadata"]["source_status"]
            }

            self.assertEqual(statuses["github_star_fallback"], "ok")
            self.assertTrue(all(
                row["total_stars"] is not None
                for row in model["rankings"]["daily"]
            ))

    def test_failed_github_top_does_not_promote_selected_set_to_total_ranking(self):
        def top_failed(top=20):
            raise SourceError("unavailable")

        with tempfile.TemporaryDirectory() as tmp:
            collected = self.app(
                tmp, good_services(fetch_top=top_failed)
            )._collect_report(report_args())
            self.assertIsNotNone(collected)
            self.assertEqual(collected[1]["rankings"]["total"], [])

    def test_local_growth_and_external_growth_remain_separate(self):
        history = [
            {
                "date": "2026-07-13",
                "at": "2026-07-13T08:30:00+00:00",
                "repo": "org/tool",
                "stars": 1900,
            },
            {
                "date": "2026-07-07",
                "at": "2026-07-07T08:30:00+00:00",
                "repo": "org/tool",
                "stars": 1500,
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / "运行状态" / "历史指标").mkdir(parents=True)
            collected = self.app(
                tmp, good_services(load_history=lambda root: copy.deepcopy(history))
            )._collect_report(report_args())
            tool = next(
                row for row in collected[1]["repositories"]
                if row["full_name"] == "org/tool"
            )
            self.assertEqual(tool["stars_24h_external"], 120)
            self.assertEqual(tool["stars_7d_external"], 640)
            self.assertEqual(tool["stars_24h_local"], 100)
            self.assertEqual(tool["stars_7d_local"], 500)

    def test_incomplete_top50_suppresses_absence_but_keeps_growth_drop(self):
        history = [
            {
                "date": "2026-07-13", "repo": "org/tool",
                "stars": 1900, "rank": 1, "stars_7d_external": 10_000,
            },
            {
                "date": "2026-07-13", "repo": "gone/project",
                "stars": 5000, "rank": 2,
            },
        ]

        def trend(period):
            if period == "24h":
                return []
            return [RepositoryRecord(
                full_name="org/tool",
                stars_7d_external=100,
                source_records=[{
                    "source": "ossinsight", "scope": "ranking", "period": "7d",
                }],
            )]

        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / "运行状态" / "历史指标").mkdir(parents=True)
            model = self.app(
                tmp,
                good_services(
                    fetch_trend=trend,
                    load_history=lambda root: copy.deepcopy(history),
                ),
            )._collect_report(report_args())[1]
            cooling = {row["repo"]: row for row in model["history"]}

            self.assertIn("org/tool", cooling)
            self.assertIn("growth_drop_70pct", cooling["org/tool"]["facts"][0])
            self.assertNotIn("fell_out", cooling["org/tool"]["facts"][0])
            self.assertNotIn("gone/project", cooling)

    def test_total_50_allows_rank_absence_even_when_weekly_is_empty(self):
        rankings = {
            "total": [
                {"repo": "org/r{:02d}".format(index), "value": 1000 - index}
                for index in range(1, 51)
            ],
            "daily": [],
            "weekly": [],
            "acceleration": [],
        }
        history = [{
            "date": "2026-07-13",
            "repo": "gone/project",
            "stars": 5000,
            "rank": 1,
        }]
        with tempfile.TemporaryDirectory() as tmp:
            result = self.app(tmp, good_services())._cooling_section(
                [], history, [], rankings
            )

        self.assertEqual([row["repo"] for row in result], ["gone/project"])
        self.assertIn("fell_out_of_top50", result[0]["facts"][0])

    def test_report_cooling_detects_only_confirmed_category_rank_drop(self):
        records = [
            RepositoryRecord(
                full_name="org/r{:02d}".format(index),
                primary_category="AI",
            )
            for index in range(1, 13)
        ]
        current_order = list(range(2, 13)) + [1]
        rankings = {
            "total": [
                {
                    "repo": "org/r{:02d}".format(index),
                    "value": 1000 - current_rank,
                }
                for current_rank, index in enumerate(current_order, 1)
            ],
            "daily": [], "weekly": [], "acceleration": [],
        }
        previous = [
            {
                "date": "2026-07-13",
                "repo": "org/r{:02d}".format(rank),
                "stars": 1000 - rank,
                "rank": rank,
                "primary_category": "AI",
            }
            for rank in range(1, 13)
        ]
        unconfirmed = copy.deepcopy(previous)
        unconfirmed[0].pop("primary_category")

        with tempfile.TemporaryDirectory() as tmp:
            app = self.app(tmp, good_services())
            confirmed = app._cooling_section(records, previous, [], rankings)
            missing_category = app._cooling_section(
                records, unconfirmed, [], rankings
            )

        self.assertTrue(any(
            row["repo"] == "org/r01"
            and "category_rank_drop" in row["facts"][0]
            for row in confirmed
        ))
        self.assertFalse(any(
            row["repo"] == "org/r01"
            and "category_rank_drop" in row["facts"][0]
            for row in missing_category
        ))

    def test_report_cooling_requires_strict_three_point_slowdown(self):
        record = RepositoryRecord(full_name="org/tool", primary_category="AI")
        history = [
            {
                "date": "2026-07-12", "repo": "org/tool", "stars": 1000,
                "rank": 1, "primary_category": "AI",
                "stars_7d_external": 300,
            },
            {
                "date": "2026-07-13", "repo": "org/tool", "stars": 1100,
                "rank": 1, "primary_category": "AI",
                "stars_7d_external": 200,
            },
            {
                "date": "2026-07-14", "repo": "org/tool", "stars": 1150,
                "rank": 1, "primary_category": "AI",
                "stars_7d_external": 999,
            },
        ]

        def analyze(current_growth, selected_history=None):
            rankings = {
                "total": [{"repo": "org/tool", "value": 1150}],
                "daily": [],
                "weekly": [{"repo": "org/tool", "value": current_growth}],
                "acceleration": [],
            }
            with tempfile.TemporaryDirectory() as tmp:
                return self.app(tmp, good_services())._cooling_section(
                    [record],
                    copy.deepcopy(selected_history or history),
                    [],
                    rankings,
                    current_date="2026-07-14",
                )

        decreasing = analyze(100)
        non_decreasing = analyze(250)
        only_two_points = analyze(100, history[1:])

        self.assertTrue(any(
            "consecutive_slowdown" in row["facts"][0]
            for row in decreasing
        ))
        self.assertFalse(any(
            "consecutive_slowdown" in row["facts"][0]
            for row in non_decreasing
        ))
        self.assertFalse(any(
            "consecutive_slowdown" in row["facts"][0]
            for row in only_two_points
        ))

    def test_unexpected_source_exception_does_not_replace_existing_latest(self):
        def crashed(top=20):
            raise RuntimeError("token=unexpected-secret")

        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            paths.latest_report.parent.mkdir(parents=True)
            paths.latest_report.write_text("old report", encoding="utf-8")
            before = paths.latest_report.read_bytes()
            code, stdout, stderr = self.capture_report(
                self.app(tmp, good_services(fetch_top=crashed)), report_args()
            )
            self.assertNotEqual(code, 0)
            self.assertEqual(paths.latest_report.read_bytes(), before)
            self.assertNotIn("unexpected-secret", stdout + stderr)

    def test_all_discovery_sources_failed_without_local_basis_does_not_save(self):
        def failed(*args, **kwargs):
            raise SourceError("token=do-not-leak")

        with tempfile.TemporaryDirectory() as tmp:
            code, stdout, stderr = self.capture_report(
                self.app(tmp, good_services(
                    fetch_top=failed,
                    fetch_trend=failed,
                    fetch_trending=failed,
                )),
                report_args(),
            )
            paths = RadarPaths(Path(tmp), "2026-07-14")
            self.assertNotEqual(code, 0)
            self.assertFalse(paths.archive_report.exists())
            self.assertFalse(paths.latest_report.exists())
            self.assertFalse(paths.history_file.exists())
            self.assertNotIn("do-not-leak", stdout + stderr)


class ReportPersistenceTests(unittest.TestCase):
    def app(self, workspace, services):
        return RadarApp(
            workspace=workspace,
            services=services,
            today=lambda: date(2026, 7, 14),
            now=lambda: datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
        )

    def capture_report(self, app, args):
        stdout, stderr = io.StringIO(), io.StringIO()
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            code = app.run_report(args)
        return code, stdout.getvalue(), stderr.getvalue()

    def test_full_report_saves_identical_fixed_archive_latest_and_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            app = self.app(tmp, good_services())
            code, stdout, stderr = self.capture_report(app, report_args())
            paths = RadarPaths(Path(tmp), "2026-07-14")
            self.assertEqual(code, 0, stderr)
            self.assertIn(str(paths.latest_report), stdout)
            self.assertEqual(
                paths.archive_report.read_bytes(), paths.latest_report.read_bytes()
            )
            self.assertEqual(
                paths.archive_data.read_bytes(), paths.latest_data.read_bytes()
            )
            self.assertTrue(paths.history_file.exists())
            initial_names = sorted(path.name for path in paths.archive_dir.iterdir())

            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()), report_args()
            )
            self.assertEqual(code, 0, stderr)
            self.assertEqual(
                sorted(path.name for path in paths.archive_dir.iterdir()),
                initial_names,
            )
            model = json.loads(paths.latest_data.read_text(encoding="utf-8"))
            self.assertEqual(model["rankings"]["daily"][0]["value"], 120)
            self.assertEqual(model["rankings"]["weekly"][0]["value"], 640)

    def test_full_report_removes_stale_latest_dated_and_alias_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            latest_dir = paths.latest_report.parent
            latest_dir.mkdir(parents=True)
            stale_md = latest_dir / "GitHub开源趋势与项目复用雷达-2026-07-13.md"
            stale_json = latest_dir / "原始数据-2026-07-13.json"
            alias_md = latest_dir / "GitHub开源趋势与项目复用雷达-最新.md"
            alias_json = latest_dir / "原始数据-最新.json"
            for path, payload in (
                (stale_md, "old md"),
                (stale_json, "{}"),
                (alias_md, "alias md"),
                (alias_json, "{}"),
            ):
                path.write_text(payload, encoding="utf-8")

            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()), report_args()
            )
            self.assertEqual(code, 0, stderr)
            self.assertTrue(paths.latest_report.exists())
            self.assertTrue(paths.latest_data.exists())
            self.assertTrue(paths.latest_html.exists())
            self.assertFalse(stale_md.exists())
            self.assertFalse(stale_json.exists())
            self.assertFalse(alias_md.exists())
            self.assertFalse(alias_json.exists())
            remaining = sorted(path.name for path in latest_dir.iterdir())
            self.assertEqual(
                remaining,
                sorted(
                    [
                        paths.latest_data.name,
                        paths.latest_html.name,
                        paths.latest_report.name,
                    ]
                ),
            )

    def test_cleanup_stale_latest_reports_keeps_only_run_date_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            latest_dir = Path(tmp)
            keep = latest_dir / "原始数据-2026-07-28.json"
            drop = latest_dir / "原始数据-2026-07-27.json"
            alias = latest_dir / "原始数据-最新.json"
            keep.write_text("{}", encoding="utf-8")
            drop.write_text("{}", encoding="utf-8")
            alias.write_text("{}", encoding="utf-8")
            cleanup_stale_latest_reports(latest_dir, "2026-07-28")
            self.assertTrue(keep.exists())
            self.assertFalse(drop.exists())
            self.assertFalse(alias.exists())


    def test_saved_runs_share_one_root_lock_for_complete_transaction(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/tool")
            first_inside = threading.Event()
            release_first = threading.Event()
            second_fetched = threading.Event()

            def services_for(stars, release, first=False):
                def top(top=20):
                    if first:
                        first_inside.set()
                        if not release_first.wait(5):
                            raise RuntimeError("test synchronization timeout")
                    else:
                        second_fetched.set()
                    return [RepositoryRecord(
                        full_name="org/tool",
                        total_stars=stars,
                        source_records=[{"source": "github", "scope": "search"}],
                    )]

                def detail(name):
                    return RepositoryRecord(
                        full_name=name,
                        total_stars=stars,
                        latest_release=release,
                        pushed_at="2026-07-14T00:00:00Z",
                        source_records=[{
                            "source": "github",
                            "scope": "repository_detail",
                            "release": "available",
                        }],
                    )

                return good_services(
                    fetch_top=top,
                    fetch_details=detail,
                    fetch_trend=lambda period: [],
                    fetch_trending=lambda period="weekly": [],
                )

            first = self.app(tmp, services_for(100, "v1", first=True))
            second = self.app(tmp, services_for(200, "v2"))
            results = {}

            def run(name, app):
                results[name] = app.run_report(report_args())

            with patch("builtins.print"):
                first_thread = threading.Thread(target=run, args=("first", first))
                second_thread = threading.Thread(target=run, args=("second", second))
                first_thread.start()
                self.assertTrue(first_inside.wait(5))
                second_thread.start()
                self.assertFalse(second_fetched.wait(0.2))
                release_first.set()
                first_thread.join(10)
                second_thread.join(10)

            self.assertFalse(first_thread.is_alive())
            self.assertFalse(second_thread.is_alive())
            self.assertEqual(results, {"first": 0, "second": 0})
            self.assertTrue(second_fetched.is_set())
            self.assertEqual(
                paths.archive_data.read_bytes(), paths.latest_data.read_bytes()
            )
            model = json.loads(paths.latest_data.read_text(encoding="utf-8"))
            self.assertEqual(model["rankings"]["total"][0]["value"], 200)
            history_rows = [
                json.loads(line)
                for line in paths.history_file.read_text(encoding="utf-8").splitlines()
            ]
            tool_history = next(
                row for row in history_rows if row["repo"] == "org/tool"
            )
            self.assertEqual(tool_history["stars"], 200)
            workbook = load_workbook(paths.watchlist_file)
            try:
                self.assertEqual(workbook[SHEET_NAME]["N2"].value, "v2")
            finally:
                workbook.close()

    def test_category_save_uses_run_lock_but_preview_does_not(self):
        calls = []

        @contextmanager
        def observed_lock(path):
            calls.append(Path(path))
            yield

        with tempfile.TemporaryDirectory() as tmp, patch(
            "radar.exclusive_file_lock", side_effect=observed_lock
        ):
            app = self.app(tmp, good_services())
            code, _, stderr = self.capture_report(
                app, report_args(category="AI")
            )
            self.assertEqual(code, 0, stderr)
            self.assertEqual(calls, [])

            code, _, stderr = self.capture_report(
                app, report_args(category="AI", save=True)
            )
            self.assertEqual(code, 0, stderr)
            self.assertEqual(
                calls,
                [Path(tmp) / "运行状态" / "完整报告运行"],
            )
            lock_target = calls[0]
            paths = RadarPaths(Path(tmp), "2026-07-14")
            self.assertNotIn(lock_target, {
                paths.archive_report,
                paths.archive_data,
                paths.latest_report,
                paths.latest_data,
                paths.history_file,
                paths.watchlist_file,
            })

    def test_shanghai_clock_is_single_source_across_midnight_run_outputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            app = RadarApp(
                workspace=tmp,
                services=good_services(),
                today=lambda: date(2026, 7, 13),
                now=lambda: datetime(
                    2026, 7, 13, 16, 30, tzinfo=timezone.utc
                ),
            )
            code, _, stderr = self.capture_report(app, report_args())
            paths = RadarPaths(Path(tmp), "2026-07-14")

            self.assertEqual(code, 0, stderr)
            model = json.loads(paths.latest_data.read_text(encoding="utf-8"))
            self.assertEqual(model["metadata"]["date"], "2026-07-14")
            self.assertEqual(
                model["metadata"]["query_time"],
                "2026-07-14T00:30:00+08:00",
            )
            history_rows = [
                json.loads(line)
                for line in paths.history_file.read_text(encoding="utf-8").splitlines()
            ]
            self.assertTrue(history_rows)
            self.assertTrue(all(row["date"] == "2026-07-14" for row in history_rows))
            self.assertTrue(all(
                row["at"] == "2026-07-14T00:30:00+08:00"
                for row in history_rows
            ))

    def test_naive_clock_fails_safely_without_creating_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            app = RadarApp(
                workspace=tmp,
                services=good_services(),
                now=lambda: datetime(2026, 7, 14, 8, 30),
            )
            code, stdout, stderr = self.capture_report(app, report_args())

            self.assertNotEqual(code, 0)
            self.assertEqual(stdout, "")
            self.assertIn("报告运行失败", stderr)
            self.assertFalse((Path(tmp) / "最新报告").exists())

    def test_watchlist_only_repository_is_isolated_until_tracking_and_history(self):
        events = []

        class OrderedRadarApp(RadarApp):
            def _objective_parts(self, *args, **kwargs):
                events.append("objective")
                return super()._objective_parts(*args, **kwargs)

        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "watch/only")
            (Path(tmp) / "运行状态" / "历史指标").mkdir(parents=True)

            base = good_services()
            original_read = base.read_watchlist
            original_detail = base.fetch_details

            def source_top(top=20):
                events.append("source")
                return good_services().fetch_top(top=top)

            def detail(name):
                events.append(
                    "watch_detail" if name == "watch/only" else "objective_detail"
                )
                return original_detail(name)

            def history(root):
                events.append("history")
                return []

            def watchlist(path):
                events.append("watchlist")
                return original_read(path)

            services = good_services(
                fetch_top=source_top,
                fetch_details=detail,
                load_history=history,
                read_watchlist=watchlist,
            )
            app = OrderedRadarApp(
                workspace=tmp,
                services=services,
                today=lambda: date(2026, 7, 14),
                now=lambda: datetime(
                    2026, 7, 14, 8, 30, tzinfo=timezone.utc
                ),
            )
            code, _, stderr = self.capture_report(app, report_args())

            self.assertEqual(code, 0, stderr)
            self.assertLess(events.index("source"), events.index("objective_detail"))
            self.assertLess(events.index("objective_detail"), events.index("history"))
            self.assertLess(events.index("history"), events.index("objective"))
            self.assertLess(events.index("objective"), events.index("watchlist"))
            self.assertLess(events.index("watchlist"), events.index("watch_detail"))

            model = json.loads(paths.latest_data.read_text(encoding="utf-8"))
            self.assertNotIn(
                "watch/only", {row["full_name"] for row in model["repositories"]}
            )
            for ranking in model["rankings"].values():
                self.assertNotIn("watch/only", {row["repo"] for row in ranking})
            self.assertNotIn(
                "watch/only",
                {row["repo"] for row in model["reusable_projects"]},
            )
            objective_counts = {}
            for repository in model["repositories"]:
                category = repository["primary_category"]
                objective_counts[category] = objective_counts.get(category, 0) + 1
            for trend in model["category_trends"]:
                self.assertEqual(
                    trend["facts"],
                    ["本期纳入 {} 个仓库。".format(objective_counts[trend["category"]])],
                )
            self.assertIn("watch/only", {row["repo"] for row in model["watchlist"]})
            self.assertIn("watch/only", {row["repo"] for row in model["history"]})

            history_rows = [
                json.loads(line)
                for line in paths.history_file.read_text(encoding="utf-8").splitlines()
            ]
            self.assertIn("watch/only", {row["repo"] for row in history_rows})

    def test_category_save_explicitly_persists_after_default_preview_does_not(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()), report_args(category="AI")
            )
            self.assertEqual(code, 0, stderr)
            self.assertFalse(paths.latest_report.exists())

            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()),
                report_args(category="AI", save=True),
            )
            self.assertEqual(code, 0, stderr)
            self.assertFalse(paths.latest_report.exists())
            self.assertTrue(
                (paths.latest_report.parent
                 / "GitHub开源趋势与项目复用雷达-分类-AI-2026-07-14.md").exists()
            )
            self.assertTrue(
                (paths.latest_data.parent
                 / "原始数据-分类-AI-2026-07-14.json").exists()
            )
            self.assertFalse(paths.history_file.exists())
            self.assertFalse(paths.watchlist_file.exists())

    def test_category_save_keeps_full_history_and_excel_bytes_unchanged(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/tool")
            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()), report_args()
            )
            self.assertEqual(code, 0, stderr)
            full_files = {
                path: path.read_bytes() for path in (
                    paths.archive_report,
                    paths.archive_data,
                    paths.latest_report,
                    paths.latest_data,
                )
            }
            history_before = paths.history_file.read_bytes()
            excel_before = paths.watchlist_file.read_bytes()

            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()),
                report_args(category="AI", save=True),
            )

            self.assertEqual(code, 0, stderr)
            for path, content in full_files.items():
                self.assertEqual(path.read_bytes(), content)
            self.assertEqual(paths.history_file.read_bytes(), history_before)
            self.assertEqual(paths.watchlist_file.read_bytes(), excel_before)
            category_archive_report = (
                paths.archive_dir
                / "GitHub开源趋势与项目复用雷达-分类-AI-2026-07-14.md"
            )
            category_archive_data = (
                paths.archive_dir / "原始数据-分类-AI-2026-07-14.json"
            )
            category_latest_report = (
                paths.latest_report.parent
                / "GitHub开源趋势与项目复用雷达-分类-AI-2026-07-14.md"
            )
            category_latest_data = (
                paths.latest_data.parent / "原始数据-分类-AI-2026-07-14.json"
            )
            self.assertEqual(
                category_archive_report.read_bytes(),
                category_latest_report.read_bytes(),
            )
            self.assertEqual(
                category_archive_data.read_bytes(),
                category_latest_data.read_bytes(),
            )
            history_rows = [
                json.loads(line) for line in history_before.decode("utf-8").splitlines()
            ]
            self.assertTrue(all(row.get("primary_category") for row in history_rows))

    def test_category_query_ignores_other_and_uncategorized_history_rows(self):
        history = [
            {
                "date": "2026-07-13",
                "at": "2026-07-13T08:30:00+00:00",
                "repo": "org/tool",
                "stars": 1900,
            },
            {
                "date": "2026-07-13",
                "repo": "legacy/non-ai",
                "stars": 5000,
                "rank": 1,
                "primary_category": "开发工具",
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / "运行状态" / "历史指标").mkdir(parents=True)
            collected = self.app(
                tmp,
                good_services(load_history=lambda root: copy.deepcopy(history)),
            )._collect_report(report_args(category="AI"))
            model = collected[1]
            tool = next(
                row for row in model["repositories"]
                if row["full_name"] == "org/tool"
            )

            self.assertIsNone(tool["stars_24h_local"])
            self.assertNotIn(
                "legacy/non-ai", {row["repo"] for row in model["history"]}
            )

    def test_archive_save_failure_skips_history_and_excel_without_secret_leak(self):
        calls = {"history": 0, "excel": 0}

        def save_failed(*args):
            raise OSError("token=hidden")

        with tempfile.TemporaryDirectory() as tmp:
            services = good_services(
                save_report=save_failed,
                upsert_history=lambda *args: calls.__setitem__("history", 1),
                update_watchlist=lambda *args: calls.__setitem__("excel", 1),
            )
            code, stdout, stderr = self.capture_report(
                self.app(tmp, services), report_args()
            )
            self.assertNotEqual(code, 0)
            self.assertEqual(calls, {"history": 0, "excel": 0})
            self.assertNotIn("token=hidden", stdout + stderr)

    def test_latest_failure_keeps_archive_and_skips_post_updates(self):
        calls = {"history": 0, "excel": 0}

        def fail_latest(report_path, data_path, markdown, model):
            if "最新" in str(report_path):
                raise OSError("secret latest failure")
            save_complete_report(report_path, data_path, markdown, model)

        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            paths.latest_report.parent.mkdir(parents=True)
            paths.latest_report.write_text("old latest report", encoding="utf-8")
            paths.latest_data.write_text("old latest data", encoding="utf-8")
            old_report = paths.latest_report.read_bytes()
            old_data = paths.latest_data.read_bytes()
            services = good_services(
                save_report=fail_latest,
                upsert_history=lambda *args: calls.__setitem__("history", 1),
                update_watchlist=lambda *args: calls.__setitem__("excel", 1),
            )
            code, stdout, stderr = self.capture_report(
                self.app(tmp, services), report_args()
            )
            self.assertNotEqual(code, 0)
            self.assertTrue(paths.archive_report.exists())
            self.assertEqual(paths.latest_report.read_bytes(), old_report)
            self.assertEqual(paths.latest_data.read_bytes(), old_data)
            self.assertEqual(calls, {"history": 0, "excel": 0})
            self.assertIn("最新报告保存失败", stderr)
            self.assertNotIn("secret latest", stdout + stderr)

    def test_history_post_failure_keeps_report_and_skips_excel(self):
        calls = []

        def history_failed(*args):
            raise ValueError("token=history-secret")

        with tempfile.TemporaryDirectory() as tmp:
            services = good_services(
                upsert_history=history_failed,
                update_watchlist=lambda *args: calls.append(args),
            )
            code, stdout, stderr = self.capture_report(
                self.app(tmp, services), report_args()
            )
            paths = RadarPaths(Path(tmp), "2026-07-14")
            self.assertNotEqual(code, 0)
            self.assertTrue(paths.latest_report.exists())
            self.assertEqual(calls, [])
            self.assertIn("报告已保存但后置更新失败", stderr)
            self.assertNotIn("history-secret", stdout + stderr)

    def test_excel_post_failure_keeps_saved_report_and_history(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/tool")

            def excel_failed(*args):
                raise ValueError("token=excel-secret")

            code, stdout, stderr = self.capture_report(
                self.app(tmp, good_services(update_watchlist=excel_failed)),
                report_args(),
            )
            self.assertNotEqual(code, 0)
            self.assertTrue(paths.latest_report.exists())
            self.assertTrue(paths.history_file.exists())
            self.assertIn("报告已保存但后置更新失败", stderr)
            self.assertNotIn("excel-secret", stdout + stderr)

    def test_successful_report_uses_repository_detail_release_for_watchlist(self):
        with tempfile.TemporaryDirectory() as tmp:
            paths = RadarPaths(Path(tmp), "2026-07-14")
            create_watchlist(paths.watchlist_file)
            append_watchlist(paths.watchlist_file, "org/tool")
            code, _, stderr = self.capture_report(
                self.app(tmp, good_services()), report_args()
            )
            self.assertEqual(code, 0, stderr)
            workbook = load_workbook(paths.watchlist_file)
            worksheet = workbook[SHEET_NAME]
            self.assertEqual(worksheet["N2"].value, "v1.2.3")
            self.assertEqual(worksheet["O2"].value, "活跃")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
