"""Excel-backed project watchlist with protected automatic fields."""

import copy
import math
import posixpath
import re
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Mapping
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, NamedTuple, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from models import RepositoryRecord
from storage import (
    StorageError,
    atomic_replace_file,
    exclusive_file_lock,
    install_new_workbook,
    secure_target_path,
)


SHEET_NAME = "项目观察清单"
MANUAL_FIELDS: Tuple[str, ...] = (
    "仓库",
    "项目名称",
    "使用状态",
    "优先级",
    "使用场景",
    "当前版本",
    "开始关注日期",
    "替代项目",
    "是否检查更新",
    "是否分析替代关系",
    "负责人",
    "备注",
)
AUTOMATIC_FIELDS: Tuple[str, ...] = (
    "最近检查时间",
    "最新版本",
    "维护状态",
    "建议动作",
)
HEADERS: Tuple[str, ...] = MANUAL_FIELDS + AUTOMATIC_FIELDS

_COLUMN_WIDTHS = (28, 22, 12, 10, 28, 14, 16, 24, 16, 20, 14, 32, 16, 16, 14, 24)
_MANUAL_FILL = PatternFill("solid", fgColor="FF4472C4")
_AUTOMATIC_FILL = PatternFill("solid", fgColor="FF70AD47")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_REFERENCE = re.compile(r"([A-Z]+)([1-9][0-9]*)")
_MAX_SAFE_INTEGER = 999_999_999_999_999
_MIN_EXCEL_FLOAT = 2.2250738585072014e-308


class _XmlTag(NamedTuple):
    start: int
    end: int
    name: bytes
    closing: bool
    self_closing: bool
    raw: bytes


class WatchlistError(ValueError):
    """Raised when a watchlist is unsafe to read or update."""


def _validate_structure(workbook):
    if SHEET_NAME not in workbook.sheetnames:
        raise WatchlistError("watchlist worksheet is missing")
    worksheet = workbook[SHEET_NAME]
    actual = tuple(worksheet.cell(1, column).value for column in range(1, 17))
    if worksheet.max_column != 16 or actual != HEADERS:
        raise WatchlistError("watchlist headers must match the fixed 16-column schema")
    return worksheet


def _add_list_validation(worksheet, column: str, choices: Tuple[str, ...]) -> None:
    validation = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(choices)),
        allow_blank=True,
    )
    worksheet.add_data_validation(validation)
    validation.add("{}2:{}1000".format(column, column))


def create_watchlist(path: Union[str, Path]) -> None:
    """Create the fixed project-watchlist template."""

    try:
        target = secure_target_path(path)
        with exclusive_file_lock(target):
            if target.exists():
                try:
                    existing = load_workbook(target, data_only=False)
                except Exception as exc:
                    raise WatchlistError("unable to open existing watchlist") from exc
                try:
                    _validate_structure(existing)
                finally:
                    existing.close()
                return

            workbook = Workbook()
            try:
                worksheet = workbook.active
                worksheet.title = SHEET_NAME
                worksheet.append(HEADERS)
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = "A1:P1"
                worksheet.row_dimensions[1].height = 24

                for index, header_cell in enumerate(worksheet[1], start=1):
                    header_cell.fill = (
                        _MANUAL_FILL if index <= 12 else _AUTOMATIC_FILL
                    )
                    header_cell.font = _HEADER_FONT
                    header_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    worksheet.column_dimensions[
                        get_column_letter(index)
                    ].width = _COLUMN_WIDTHS[index - 1]

                _add_list_validation(
                    worksheet, "C", ("未使用", "试用中", "使用中", "已停用")
                )
                _add_list_validation(worksheet, "D", ("高", "中", "低"))
                _add_list_validation(worksheet, "I", ("是", "否"))
                _add_list_validation(worksheet, "J", ("是", "否"))
                install_new_workbook(workbook, target, _validate_structure)
            finally:
                workbook.close()
    except StorageError as exc:
        raise WatchlistError("unable to create watchlist safely") from exc


def _normalized_repository(value: Any) -> str:
    if not isinstance(value, str):
        raise WatchlistError("repository must use owner/repo format")
    try:
        return RepositoryRecord(full_name=value).full_name
    except (TypeError, ValueError, AttributeError) as exc:
        raise WatchlistError("repository must use owner/repo format") from exc


def _read_rows(worksheet) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    repositories = set()
    for row_number in range(2, worksheet.max_row + 1):
        cells = [worksheet.cell(row_number, column) for column in range(1, 17)]
        if all(cell.value is None or cell.value == "" for cell in cells):
            continue
        if any(cell.data_type in {"f", "e"} for cell in cells):
            raise WatchlistError("watchlist contains an unsafe cell value")

        repository = _normalized_repository(cells[0].value)
        if repository in repositories:
            raise WatchlistError("watchlist contains a duplicate repository")
        repositories.add(repository)

        values = [cell.value for cell in cells]
        values[0] = repository
        rows.append(dict(zip(HEADERS, values)))
    return rows


def read_watchlist(path: Union[str, Path]) -> List[Dict[str, Any]]:
    """Read validated watchlist rows, creating the template when absent."""

    try:
        target = secure_target_path(path)
    except StorageError as exc:
        raise WatchlistError("unable to read watchlist safely") from exc
    if not target.exists():
        create_watchlist(target)
    try:
        workbook = load_workbook(target, data_only=False)
    except Exception as exc:
        raise WatchlistError("unable to open watchlist") from exc
    try:
        worksheet = _validate_structure(workbook)
        return _read_rows(worksheet)
    finally:
        workbook.close()


def _valid_xml_text(value: str) -> bool:
    for character in value:
        codepoint = ord(character)
        if codepoint in (0x09, 0x0A, 0x0D):
            continue
        if 0x20 <= codepoint <= 0xD7FF:
            continue
        if 0xE000 <= codepoint <= 0xFFFD:
            continue
        if 0x10000 <= codepoint <= 0x10FFFF:
            continue
        return False
    return True


def _safe_automatic_value(value: Any) -> Any:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            raise WatchlistError("automatic field value must be a safe scalar")
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, int):
        if abs(value) <= _MAX_SAFE_INTEGER:
            return value
        raise WatchlistError("automatic field value must be a safe scalar")
    if isinstance(value, float):
        if math.isfinite(value) and (
            value == 0.0 or abs(value) >= _MIN_EXCEL_FLOAT
        ):
            return value
        raise WatchlistError("automatic field value must be a safe scalar")
    if isinstance(value, str):
        if len(value) > 32767 or not _valid_xml_text(value):
            raise WatchlistError("automatic field value must be safe XML text")
        if value.lstrip().startswith(("=", "+", "-", "@")):
            raise WatchlistError("formula-like automatic field values are not allowed")
        return value
    raise WatchlistError("automatic field value must be a safe scalar")


def _validated_updates(updates: Any) -> Dict[str, Dict[str, Any]]:
    if not isinstance(updates, Mapping):
        raise WatchlistError("updates must map repositories to automatic fields")
    validated: Dict[str, Dict[str, Any]] = {}
    for repository_value, field_values in updates.items():
        repository = _normalized_repository(repository_value)
        if repository in validated:
            raise WatchlistError("updates contain a duplicate repository")
        if not isinstance(field_values, Mapping):
            raise WatchlistError("repository update must be a field mapping")
        unknown = set(field_values) - set(AUTOMATIC_FIELDS)
        if unknown:
            raise WatchlistError("updates may contain only automatic fields")
        normalized_fields = {
            field: _safe_automatic_value(value)
            for field, value in field_values.items()
        }
        normalized_fields.setdefault("最近检查时间", date.today().isoformat())
        validated[repository] = normalized_fields
    return validated


def _automatic_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _update_repository_rows(worksheet) -> Dict[str, int]:
    rows: Dict[str, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        cells = [worksheet.cell(row_number, column) for column in range(1, 17)]
        if all(cell.value is None or cell.value == "" for cell in cells):
            continue
        repository_cell = cells[0]
        if repository_cell.data_type in {"f", "e"}:
            raise WatchlistError("watchlist contains an unsafe repository cell")
        if any(cell.data_type in {"f", "e"} for cell in cells[12:]):
            raise WatchlistError("watchlist contains an unsafe automatic cell")
        repository = _normalized_repository(repository_cell.value)
        if repository in rows:
            raise WatchlistError("watchlist contains a duplicate repository")
        rows[repository] = row_number
    return rows


def _qualified(namespace: str, name: str) -> str:
    return "{{{}}}{}".format(namespace, name)


def _worksheet_archive_path(archive: zipfile.ZipFile) -> str:
    try:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ET.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
    except (KeyError, ET.ParseError) as exc:
        raise WatchlistError("workbook package structure is invalid") from exc

    relationship_id = None
    for sheet in workbook_root.findall(
        ".//{}".format(_qualified(_MAIN_NS, "sheet"))
    ):
        if sheet.get("name") == SHEET_NAME:
            relationship_id = sheet.get(_qualified(_DOC_REL_NS, "id"))
            break
    if not relationship_id:
        raise WatchlistError("watchlist worksheet relationship is missing")

    relationship_target = None
    for relationship in relationships_root.findall(
        _qualified(_PACKAGE_REL_NS, "Relationship")
    ):
        if relationship.get("Id") == relationship_id:
            if relationship.get("TargetMode") == "External":
                raise WatchlistError("watchlist worksheet relationship is invalid")
            relationship_target = relationship.get("Target")
            break
    if not relationship_target:
        raise WatchlistError("watchlist worksheet target is missing")

    normalized_target = relationship_target.replace("\\", "/")
    if normalized_target.startswith("/"):
        archive_path = posixpath.normpath(normalized_target.lstrip("/"))
    else:
        archive_path = posixpath.normpath(
            posixpath.join("xl", normalized_target)
        )
    if archive_path == ".." or archive_path.startswith("../"):
        raise WatchlistError("watchlist worksheet target is invalid")
    if archive.namelist().count(archive_path) != 1:
        raise WatchlistError("watchlist worksheet package entry is invalid")
    return archive_path


def _column_number(reference: str) -> int:
    matched = _CELL_REFERENCE.fullmatch(reference.upper())
    if matched is None:
        raise WatchlistError("worksheet contains an invalid cell reference")
    result = 0
    for character in matched.group(1):
        result = result * 26 + ord(character) - ord("A") + 1
    return result


def _tag_end(xml_bytes: bytes, start: int) -> int:
    quote = None
    index = start + 1
    while index < len(xml_bytes):
        character = xml_bytes[index]
        if quote is not None:
            if character == quote:
                quote = None
        elif character in (ord('"'), ord("'")):
            quote = character
        elif character == ord(">"):
            return index + 1
        index += 1
    raise WatchlistError("watchlist worksheet contains an incomplete XML tag")


def _scan_xml_tags(xml_bytes: bytes) -> List[_XmlTag]:
    tags: List[_XmlTag] = []
    position = 0
    while True:
        start = xml_bytes.find(b"<", position)
        if start < 0:
            break
        if xml_bytes.startswith(b"<!--", start):
            end = xml_bytes.find(b"-->", start + 4)
            if end < 0:
                raise WatchlistError("watchlist worksheet XML comment is incomplete")
            position = end + 3
            continue
        if xml_bytes.startswith(b"<?", start):
            end = xml_bytes.find(b"?>", start + 2)
            if end < 0:
                raise WatchlistError(
                    "watchlist worksheet processing instruction is incomplete"
                )
            position = end + 2
            continue
        if xml_bytes.startswith(b"<![CDATA[", start):
            end = xml_bytes.find(b"]]>", start + 9)
            if end < 0:
                raise WatchlistError("watchlist worksheet CDATA is incomplete")
            position = end + 3
            continue
        end = _tag_end(xml_bytes, start)
        raw = xml_bytes[start:end]
        inner = raw[1:-1].strip()
        if inner.startswith(b"!"):
            position = end
            continue
        closing = inner.startswith(b"/")
        if closing:
            inner = inner[1:].lstrip()
        self_closing = not closing and inner.rstrip().endswith(b"/")
        if self_closing:
            inner = inner.rstrip()[:-1].rstrip()
        name_end = 0
        while name_end < len(inner) and inner[name_end] not in b" \t\r\n/>":
            name_end += 1
        if name_end == 0:
            raise WatchlistError("watchlist worksheet contains an invalid XML tag")
        tags.append(
            _XmlTag(start, end, inner[:name_end], closing, self_closing, raw)
        )
        position = end
    return tags


def _parse_attributes(tag: _XmlTag) -> List[Tuple[bytes, bytes, bytes]]:
    raw = tag.raw
    position = 1 + len(tag.name)
    limit = len(raw) - (2 if tag.self_closing else 1)
    attributes: List[Tuple[bytes, bytes, bytes]] = []
    while position < limit:
        while position < limit and raw[position] in b" \t\r\n":
            position += 1
        if position >= limit:
            break
        name_start = position
        while position < limit and raw[position] not in b"= \t\r\n":
            position += 1
        name = raw[name_start:position]
        while position < limit and raw[position] in b" \t\r\n":
            position += 1
        if not name or position >= limit or raw[position] != ord("="):
            raise WatchlistError("watchlist XML attribute is invalid")
        position += 1
        while position < limit and raw[position] in b" \t\r\n":
            position += 1
        if position >= limit or raw[position] not in (ord('"'), ord("'")):
            raise WatchlistError("watchlist XML attribute is invalid")
        quote = bytes((raw[position],))
        position += 1
        value_start = position
        value_end = raw.find(quote, value_start, limit)
        if value_end < 0:
            raise WatchlistError("watchlist XML attribute is incomplete")
        attributes.append((name, quote, raw[value_start:value_end]))
        position = value_end + 1
    return attributes


def _attribute_value(tag: _XmlTag, name: bytes) -> bytes:
    matches = [
        value
        for attribute, _quote, value in _parse_attributes(tag)
        if attribute == name
    ]
    if len(matches) != 1:
        raise WatchlistError("watchlist XML requires one cell or row reference")
    return matches[0]


def _local_name(name: bytes) -> bytes:
    return name.rsplit(b":", 1)[-1]


def _matching_close(tags: List[_XmlTag], start_index: int) -> _XmlTag:
    opening = tags[start_index]
    if opening.self_closing:
        return opening
    depth = 0
    for candidate in tags[start_index + 1 :]:
        if candidate.name != opening.name:
            continue
        if candidate.closing:
            if depth == 0:
                return candidate
            depth -= 1
        elif not candidate.self_closing:
            depth += 1
    raise WatchlistError("watchlist worksheet XML element is incomplete")


def _element_bounds(tags: List[_XmlTag], start_index: int) -> Tuple[int, int]:
    opening = tags[start_index]
    closing = _matching_close(tags, start_index)
    if opening.self_closing:
        return opening.start, opening.end
    return opening.start, closing.end


def _sheet_data_bounds(tags: List[_XmlTag]) -> Tuple[int, int]:
    candidates = [
        index
        for index, tag in enumerate(tags)
        if not tag.closing and _local_name(tag.name) == b"sheetData"
    ]
    if len(candidates) != 1:
        raise WatchlistError("watchlist worksheet requires one sheetData element")
    opening = tags[candidates[0]]
    closing = _matching_close(tags, candidates[0])
    return opening.end, closing.start


def _row_bounds(
    tags: List[_XmlTag], row_number: int, sheet_data: Tuple[int, int]
) -> Tuple[int, int, int, int, bytes]:
    expected = str(row_number).encode("ascii")
    matches = []
    for index, tag in enumerate(tags):
        if tag.closing or _local_name(tag.name) != b"row":
            continue
        if not (sheet_data[0] <= tag.start < sheet_data[1]):
            continue
        if _attribute_value(tag, b"r") == expected:
            matches.append(index)
    if len(matches) != 1:
        raise WatchlistError("watchlist row is missing or duplicated")
    opening = tags[matches[0]]
    closing = _matching_close(tags, matches[0])
    if opening.self_closing:
        raise WatchlistError("watchlist repository row cannot be empty")
    return opening.start, opening.end, closing.start, closing.end, opening.name


def _cell_qname(row_qname: bytes) -> bytes:
    if b":" not in row_qname:
        return b"c"
    return row_qname.rsplit(b":", 1)[0] + b":c"


def _cell_content_qname(cell_qname: bytes, local: bytes) -> bytes:
    if b":" not in cell_qname:
        return local
    return cell_qname.rsplit(b":", 1)[0] + b":" + local


def _escaped_text(text: str) -> bytes:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\r", "&#13;")
        .encode("utf-8")
    )


def _inline_content(cell_qname: bytes, text: str) -> bytes:
    inline_qname = _cell_content_qname(cell_qname, b"is")
    text_qname = _cell_content_qname(cell_qname, b"t")
    space = b' xml:space="preserve"' if text != text.strip() else b""
    return (
        b"<"
        + inline_qname
        + b"><"
        + text_qname
        + space
        + b">"
        + _escaped_text(text)
        + b"</"
        + text_qname
        + b"></"
        + inline_qname
        + b">"
    )


def _new_cell(cell_qname: bytes, reference: bytes, text: str) -> bytes:
    return (
        b"<"
        + cell_qname
        + b' r="'
        + reference
        + b'" t="inlineStr">'
        + _inline_content(cell_qname, text)
        + b"</"
        + cell_qname
        + b">"
    )


def _replacement_cell(tag: _XmlTag, reference: bytes, text: str) -> bytes:
    attributes = _parse_attributes(tag)
    if _attribute_value(tag, b"r") != reference:
        raise WatchlistError("watchlist cell reference is inconsistent")
    pieces = [b"<" + tag.name]
    for name, quote, value in attributes:
        if name == b"t":
            continue
        pieces.extend((b" ", name, b"=", quote, value, quote))
    pieces.append(b' t="inlineStr">')
    pieces.append(_inline_content(tag.name, text))
    pieces.extend((b"</", tag.name, b">"))
    return b"".join(pieces)


def _apply_patches(
    xml_bytes: bytes, patches: List[Tuple[int, int, bytes]]
) -> bytes:
    ordered = sorted(patches, key=lambda item: (item[0], item[1] != item[0]))
    result = []
    cursor = 0
    for start, end, replacement in ordered:
        if start < cursor or end < start:
            raise WatchlistError("watchlist XML patches overlap")
        result.append(xml_bytes[cursor:start])
        result.append(replacement)
        cursor = end
    result.append(xml_bytes[cursor:])
    return b"".join(result)


def _transform_worksheet_xml(
    xml_bytes: bytes, updates_by_row: Dict[int, Dict[int, str]]
) -> bytes:
    tags = _scan_xml_tags(xml_bytes)
    sheet_data = _sheet_data_bounds(tags)
    patches: List[Tuple[int, int, bytes]] = []
    for row_number, column_updates in sorted(updates_by_row.items()):
        _row_start, row_content_start, row_content_end, _row_end, row_qname = (
            _row_bounds(tags, row_number, sheet_data)
        )
        expected_cell_qname = _cell_qname(row_qname)
        cells = []
        for index, tag in enumerate(tags):
            if tag.closing or tag.name != expected_cell_qname:
                continue
            if not (row_content_start <= tag.start < row_content_end):
                continue
            reference = _attribute_value(tag, b"r")
            reference_text = reference.decode("ascii", errors="strict")
            column = _column_number(reference_text)
            start, end = _element_bounds(tags, index)
            cells.append((column, reference, tag, start, end))
        columns = [cell[0] for cell in cells]
        if columns != sorted(columns) or len(columns) != len(set(columns)):
            raise WatchlistError("watchlist cells must have unique column order")
        by_column = {cell[0]: cell for cell in cells}
        insertions: Dict[int, List[Tuple[int, bytes]]] = {}
        for column, text in sorted(column_updates.items()):
            reference = "{}{}".format(
                get_column_letter(column), row_number
            ).encode("ascii")
            existing = by_column.get(column)
            if existing is not None:
                _existing_column, _reference, tag, start, end = existing
                patches.append(
                    (start, end, _replacement_cell(tag, reference, text))
                )
                continue
            later = [cell for cell in cells if cell[0] > column]
            insertion_point = (
                later[0][3]
                if later
                else (cells[-1][4] if cells else row_content_start)
            )
            insertions.setdefault(insertion_point, []).append(
                (column, _new_cell(expected_cell_qname, reference, text))
            )
        for insertion_point, values in insertions.items():
            combined = b"".join(value for _column, value in sorted(values))
            patches.append((insertion_point, insertion_point, combined))
    return _apply_patches(xml_bytes, patches)


def _copy_ooxml_with_updates(
    source_path: Path,
    temp_path: Path,
    updates_by_row: Dict[int, Dict[int, str]],
) -> None:
    try:
        with zipfile.ZipFile(source_path, "r") as source:
            worksheet_path = _worksheet_archive_path(source)
            with zipfile.ZipFile(temp_path, "w") as destination:
                destination.comment = source.comment
                for info in source.infolist():
                    with source.open(info, "r") as handle:
                        payload = handle.read()
                    if info.filename == worksheet_path:
                        payload = _transform_worksheet_xml(payload, updates_by_row)
                    destination.writestr(info, payload)
    except (OSError, zipfile.BadZipFile, RuntimeError) as exc:
        raise WatchlistError("unable to transform watchlist package safely") from exc


def update_automatic_fields(path: Union[str, Path], updates: Any) -> int:
    """Update only columns M-P under a lock; unknown repositories are ignored."""

    validated = _validated_updates(updates)
    try:
        target = secure_target_path(path)
    except StorageError as exc:
        raise WatchlistError("unable to update watchlist safely") from exc
    if not target.exists():
        create_watchlist(target)

    try:
        with exclusive_file_lock(target):
            try:
                workbook = load_workbook(target, data_only=False)
            except Exception as exc:
                raise WatchlistError("unable to open watchlist") from exc
            try:
                worksheet = _validate_structure(workbook)
                repository_rows = _update_repository_rows(worksheet)
                matched = {
                    repository: repository_rows[repository]
                    for repository in validated
                    if repository in repository_rows
                }
                if not matched:
                    return 0
                manual_snapshot = [
                    [
                        copy.copy(worksheet.cell(row, column).value)
                        for column in range(1, 13)
                    ]
                    for row in range(1, worksheet.max_row + 1)
                ]
                original_max_row = worksheet.max_row
            finally:
                workbook.close()

            field_columns = {
                field: HEADERS.index(field) + 1 for field in AUTOMATIC_FIELDS
            }
            updates_by_row: Dict[int, Dict[int, str]] = {}
            expected: Dict[Tuple[int, int], str] = {}
            for repository, row_number in matched.items():
                row_updates = updates_by_row.setdefault(row_number, {})
                for field, value in validated[repository].items():
                    column = field_columns[field]
                    text_value = _automatic_text(value)
                    row_updates[column] = text_value
                    expected[(row_number, column)] = text_value

            def writer(temp_path: Path) -> None:
                _copy_ooxml_with_updates(target, temp_path, updates_by_row)

            def validate_saved(temp_path: Path) -> None:
                try:
                    saved_workbook = load_workbook(temp_path, data_only=False)
                except Exception as exc:
                    raise WatchlistError("unable to verify updated watchlist") from exc
                try:
                    saved_sheet = _validate_structure(saved_workbook)
                    if saved_sheet.max_row != original_max_row:
                        raise WatchlistError(
                            "watchlist row structure changed during save"
                        )
                    saved_manual = [
                        [
                            saved_sheet.cell(row, column).value
                            for column in range(1, 13)
                        ]
                        for row in range(1, saved_sheet.max_row + 1)
                    ]
                    if saved_manual != manual_snapshot:
                        raise WatchlistError("manual fields changed during save")
                    _update_repository_rows(saved_sheet)
                    for (row_number, column), text_value in expected.items():
                        actual = saved_sheet.cell(row_number, column).value
                        if not (actual == text_value or (text_value == "" and actual is None)):
                            raise WatchlistError(
                                "automatic field verification failed"
                            )
                finally:
                    saved_workbook.close()

            atomic_replace_file(target, writer, validate_saved)
            return len(matched)
    except StorageError as exc:
        raise WatchlistError("unable to update watchlist safely") from exc
